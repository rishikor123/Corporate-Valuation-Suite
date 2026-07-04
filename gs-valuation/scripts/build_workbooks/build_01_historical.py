import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM1 = '0.0x;(0.0x);"-"'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=7):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']
    c.value = text
    c.font = TITLE
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = Workbook()

# ============================= COVER ===================================
ws = wb.active
ws.title = 'Cover'
title_block(ws, 'The Goldman Sachs Group, Inc. (NYSE: GS) — Historical Financial Analysis', span=6)
ws.column_dimensions['A'].width = 34
for col in 'BCDEF':
    ws.column_dimensions[col].width = 18

rows = [
    ('Prepared for', 'Corporate Valuation Suite — Portfolio Project', None),
    ('Sector', 'Financials — Global Investment Bank / Diversified Financial Services', None),
    ('Ticker / Exchange', 'GS / NYSE', None),
    ('Data as-of date', AS_OF_DATE, None),
    ('', '', None),
]
r = 3
for label, val, _ in rows:
    ws.cell(row=r, column=1, value=label).font = BOLD
    ws.cell(row=r, column=2, value=val).font = BLACK
    r += 1

r += 1
ws.cell(row=r, column=1, value='Current Snapshot').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
snapshot = [
    ('Share price', GS_PRICE, USD2),
    ('52-week range', f'${GS_52WK_LOW:,.2f} - ${GS_52WK_HIGH:,.2f}', None),
    ('Diluted shares outstanding (mm)', GS_SHARES_CURRENT_MM, '#,##0.0'),
    ('Market capitalization ($mm)', GS_PRICE * GS_SHARES_CURRENT_MM, USD0),
    ('FY2025 Book value / share', GS_HISTORY[2025]['bvps'], USD2),
    ('Price / Book (current)', GS_PRICE / GS_HISTORY[2025]['bvps'], NUM1),
    ('FY2025 Diluted EPS', GS_HISTORY[2025]['diluted_eps'], USD2),
    ('Price / FY25 EPS', GS_PRICE / GS_HISTORY[2025]['diluted_eps'], NUM1),
    ('Quarterly dividend', GS_DIV_QUARTERLY, USD2),
    ('Indicated annual dividend', GS_DIV_QUARTERLY * 4, USD2),
    ('Dividend yield', (GS_DIV_QUARTERLY * 4) / GS_PRICE, PCT1),
    ('Beta (5Y monthly, approx.)', GS_BETA, '0.00'),
]
for label, val, fmt in snapshot:
    ws.cell(row=r, column=1, value=label).font = BLACK
    cell = ws.cell(row=r, column=2, value=val)
    cell.font = BLUE if not isinstance(val, str) else BLACK
    if fmt:
        cell.number_format = fmt
    r += 1

r += 1
ws.cell(row=r, column=1, value='Contents of this workbook').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
for sheet_name, desc in [
    ('Income Statement', 'FY2021-FY2025 net revenues, net earnings, EPS, YoY growth'),
    ('Key Ratios', 'ROE, ROTE, efficiency ratio, EPS growth, margin trends'),
    ('Per Share & Capital', 'BVPS, dividends, buybacks, shares outstanding walk'),
    ('Segment Analysis', 'Net revenues by business segment, FY2021-FY2025'),
    ('Data Sources', 'Full citation list for every hardcoded input'),
]:
    ws.cell(row=r, column=1, value=sheet_name).font = BLACK
    ws.cell(row=r, column=2, value=desc).font = NOTE
    r += 1

r += 2
note = ws.cell(row=r, column=1, value=(
    'Note: Market-data fields (price, shares, market cap) are a snapshot as of the as-of date above, '
    'compiled from public secondary sources (see Data Sources tab). Secondary aggregators disagree by '
    'low single digits depending on exact pull time; refresh from a live terminal before using for '
    'live decisions or interviews.'))
note.font = NOTE
ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=6)
note.alignment = Alignment(wrap_text=True, vertical='top')

# ============================= INCOME STATEMENT ===================================
ws = wb.create_sheet('Income Statement')
title_block(ws, 'Historical Income Statement — Summary ($ in millions, except per share)', span=7)
ws.column_dimensions['A'].width = 34
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 14

years = [2021, 2022, 2023, 2024, 2025]
ws.cell(row=3, column=1, value='($ in millions, except per share data)').font = NOTE
ws.cell(row=4, column=1, value='Metric')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}A')
style_header_row(ws, 4, len(years) + 1)

r = 5
rowmap = {}

def write_row(label, values, fmt=USD0, bold=False, is_formula=False, hardcode=True):
    global r
    ws.cell(row=r, column=1, value=label).font = BOLD if bold else BLACK
    for i, v in enumerate(values):
        cell = ws.cell(row=r, column=2 + i)
        cell.value = v
        cell.number_format = fmt
        if is_formula:
            cell.font = BLACK
        elif hardcode:
            cell.font = BLUE
        cell.border = BORDER
    rowmap[label] = r
    r += 1
    return r - 1

write_row('Net revenues', [GS_HISTORY[y]['net_revenues'] for y in years])
write_row('  % growth YoY', ['n/a'] + [f'=B5*0/B5' for _ in years[1:]], fmt=PCT1, is_formula=False)
# overwrite growth row with real formulas referencing prior column
row_growth = rowmap['  % growth YoY']
for i in range(1, len(years)):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    ws.cell(row=row_growth, column=2 + i, value=f'=({col}5-{prev}5)/{prev}5')
ws.cell(row=row_growth, column=2, value='n/a')

write_row('Operating expenses', [GS_HISTORY[y]['opex'] if GS_HISTORY[y]['opex'] else '=' for y in years])
row_opex = rowmap['Operating expenses']
for i, y in enumerate(years):
    if GS_HISTORY[y]['opex'] is None:
        col = get_column_letter(2 + i)
        ws.cell(row=row_opex, column=2 + i, value=f'=NA()')
        ws.cell(row=row_opex, column=2 + i).value = None
        ws.cell(row=row_opex, column=2 + i, value='n/d')
        ws.cell(row=row_opex, column=2 + i).font = NOTE

write_row('Pre-tax / operating earnings (Rev - Opex)', ['=' for _ in years], is_formula=True)
row_pretax = rowmap['Pre-tax / operating earnings (Rev - Opex)']
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    if GS_HISTORY[y]['opex'] is not None:
        ws.cell(row=row_pretax, column=2 + i, value=f'={col}5-{col}{row_opex}')
    else:
        ws.cell(row=row_pretax, column=2 + i, value='n/d')
        ws.cell(row=row_pretax, column=2 + i).font = NOTE

write_row('Net earnings', [GS_HISTORY[y]['net_earnings'] for y in years])
write_row('Less: preferred dividends', [-GS_PREFERRED_DIV[y] for y in years])
row_ne = rowmap['Net earnings']
row_pref = rowmap['Less: preferred dividends']
write_row('Net earnings applicable to common', ['=' for _ in years], is_formula=True, bold=True)
row_common = rowmap['Net earnings applicable to common']
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    ws.cell(row=row_common, column=2 + i, value=f'={col}{row_ne}+{col}{row_pref}')

write_row('Diluted weighted-average shares (mm)', [GS_SHARES_HISTORY[y] for y in years], fmt='#,##0.0')
row_shares = rowmap['Diluted weighted-average shares (mm)']
write_row('Diluted EPS (calculated check)', ['=' for _ in years], fmt=USD2, is_formula=True)
row_eps_calc = rowmap['Diluted EPS (calculated check)']
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    ws.cell(row=row_eps_calc, column=2 + i, value=f'={col}{row_common}/{col}{row_shares}')

write_row('Diluted EPS (as reported)', [GS_HISTORY[y]['diluted_eps'] for y in years], fmt=USD2, bold=True)
row_eps_rep = rowmap['Diluted EPS (as reported)']

write_row('  EPS growth YoY', ['n/a'] + [None]*4, fmt=PCT1)
row_epsg = rowmap['  EPS growth YoY']
for i in range(1, len(years)):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    ws.cell(row=row_epsg, column=2 + i, value=f'=({col}{row_eps_rep}-{prev}{row_eps_rep})/{prev}{row_eps_rep}')

write_row('Net margin (Net earnings / Net revenues)', ['=' for _ in years], fmt=PCT1, is_formula=True)
row_margin = rowmap['Net margin (Net earnings / Net revenues)']
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    ws.cell(row=row_margin, column=2 + i, value=f'={col}{row_ne}/{col}5')

r += 1
ws.cell(row=r, column=1, value='EPS calc-vs-reported check should equal reported EPS within rounding '
        '(diluted share count is a weighted-average estimate; small variance expected).').font = NOTE
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

# CAGR box
r += 2
ws.cell(row=r, column=1, value="5-Year CAGRs (2021A-2025A)").font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
ws.cell(row=r, column=1, value='Net revenues CAGR').font = BLACK
ws.cell(row=r, column=2, value=f'=(F5/B5)^(1/4)-1').number_format = PCT1
r += 1
ws.cell(row=r, column=1, value='Diluted EPS CAGR').font = BLACK
ws.cell(row=r, column=2, value=f'=(F{row_eps_rep}/B{row_eps_rep})^(1/4)-1').number_format = PCT1

for row in ws.iter_rows(min_row=4, max_row=r, min_col=1, max_col=6):
    for cell in row:
        cell.border = BORDER

wb.save('/home/claude/gs-valuation/models/01_Historical_Financial_Analysis.xlsx')
print("Saved sheet 1 of 5 (Cover, Income Statement) - continuing in part 2 script")
