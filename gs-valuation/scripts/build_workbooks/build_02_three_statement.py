import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BOLDBLUE = Font(name='Arial', size=11, bold=True, color='0000FF')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
YELLOW = PatternFill('solid', fgColor='FFFF00')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM1 = '#,##0.0'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = Workbook()
years = PROJECTION_YEARS  # [2026..2030]
n = len(years)

# ============================= ASSUMPTIONS ===================================
ws = wb.active
ws.title = 'Assumptions'
title_block(ws, 'Integrated Model Assumptions — Bear / Base / Bull Drivers, FY2026E-FY2030E', span=8)
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 14
for col in 'CDEFG':
    ws.column_dimensions[col].width = 11

# Scenario selector
ws['A3'] = 'ACTIVE SCENARIO (type Bear / Base / Bull):'
ws['A3'].font = BOLD
ws['B3'] = 'Base'
ws['B3'].font = BOLDBLUE
ws['B3'].fill = YELLOW
ws['B3'].border = BORDER
dv = DataValidation(type='list', formula1='"Bear,Base,Bull"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws['B3'])

r0 = 5
for scen in ['Bear', 'Base', 'Bull']:
    ws.cell(row=r0, column=1, value=f'{scen} Case drivers').font = SUBHDR
    ws.cell(row=r0, column=1).fill = SUBHDR_FILL
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=1 + n)
    r0 += 1
    ws.cell(row=r0, column=1, value='Fiscal Year')
    for i, y in enumerate(years):
        ws.cell(row=r0, column=2 + i, value=f'{y}E')
    style_header_row(ws, r0, n + 1)
    r0 += 1
    s = SCENARIOS[scen]
    rowlabels = [
        ('Net revenue growth %', s['rev_growth'], PCT1),
        ('Efficiency ratio (Opex / Revenue)', s['efficiency_ratio'], PCT1),
    ]
    for label, vals, fmt in rowlabels:
        ws.cell(row=r0, column=1, value=label).font = BLACK
        for i, v in enumerate(vals):
            c = ws.cell(row=r0, column=2 + i, value=v); c.font = BLUE; c.number_format = fmt; c.border = BORDER
        r0 += 1
    single_vals = [
        ('Effective tax rate', s['tax_rate'], PCT1),
        ('Total payout ratio (div + buyback / NI to common)', s['payout_ratio'], PCT1),
        ('Balance sheet / RWA growth rate (equity retention driver)', s['bs_growth'], PCT1),
        ('Terminal growth rate (Gordon growth, post-2030E)', s['terminal_growth'], PCT2),
        ('Cost of equity (CAPM)', s['cost_of_equity'], PCT2),
    ]
    for label, v, fmt in single_vals:
        ws.cell(row=r0, column=1, value=label).font = BLACK
        c = ws.cell(row=r0, column=2, value=v); c.font = BLUE; c.number_format = fmt; c.border = BORDER
        r0 += 1
    r0 += 1

BEAR_START = 6
BASE_START = 6 + 1 + 7 + 1  # header row(1)+ 2 growth rows + 5 single rows + spacer... compute exactly below
# Recompute exact block starts by scanning is safer; instead hardcode based on construction above:
# Each block: 1 subheader + 1 yearheader + 2 (growth,eff rows) + 5 single rows + 1 spacer = 10 rows
BLOCK = 10
BEAR_HDR = 5      # 'Bear Case drivers' subheader row
BEAR_YR  = 6
BEAR_GROWTH = 7
BEAR_EFF = 8
BEAR_TAX = 9
BEAR_PAYOUT = 10
BEAR_BS = 11
BEAR_TG = 12
BEAR_KE = 13

BASE_HDR = BEAR_HDR + BLOCK
BASE_GROWTH = BEAR_GROWTH + BLOCK
BASE_EFF = BEAR_EFF + BLOCK
BASE_TAX = BEAR_TAX + BLOCK
BASE_PAYOUT = BEAR_PAYOUT + BLOCK
BASE_BS = BEAR_BS + BLOCK
BASE_TG = BEAR_TG + BLOCK
BASE_KE = BEAR_KE + BLOCK

BULL_HDR = BASE_HDR + BLOCK
BULL_GROWTH = BASE_GROWTH + BLOCK
BULL_EFF = BASE_EFF + BLOCK
BULL_TAX = BASE_TAX + BLOCK
BULL_PAYOUT = BASE_PAYOUT + BLOCK
BULL_BS = BASE_BS + BLOCK
BULL_TG = BASE_TG + BLOCK
BULL_KE = BASE_KE + BLOCK

# ---- ACTIVE DRIVERS block (formula-driven off B3 selector) ----
r0 += 1
ws.cell(row=r0, column=1, value='ACTIVE DRIVERS (auto-selected from B3)').font = SUBHDR
ws.cell(row=r0, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=1 + n)
ACTIVE_HDR = r0
r0 += 1
ws.cell(row=r0, column=1, value='Fiscal Year')
for i, y in enumerate(years):
    ws.cell(row=r0, column=2 + i, value=f'{y}E')
style_header_row(ws, r0, n + 1)
r0 += 1
ACT_GROWTH = r0
ws.cell(row=r0, column=1, value='Net revenue growth %').font = BOLD
for i in range(n):
    col = get_column_letter(2 + i)
    f = (f'=IF($B$3="Bear",{col}{BEAR_GROWTH},IF($B$3="Base",{col}{BASE_GROWTH},{col}{BULL_GROWTH}))')
    c = ws.cell(row=r0, column=2 + i, value=f); c.number_format = PCT1; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_EFF = r0
ws.cell(row=r0, column=1, value='Efficiency ratio').font = BOLD
for i in range(n):
    col = get_column_letter(2 + i)
    f = (f'=IF($B$3="Bear",{col}{BEAR_EFF},IF($B$3="Base",{col}{BASE_EFF},{col}{BULL_EFF}))')
    c = ws.cell(row=r0, column=2 + i, value=f); c.number_format = PCT1; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_TAX = r0
ws.cell(row=r0, column=1, value='Effective tax rate').font = BOLD
f = f'=IF($B$3="Bear",B{BEAR_TAX},IF($B$3="Base",B{BASE_TAX},B{BULL_TAX}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_PAYOUT = r0
ws.cell(row=r0, column=1, value='Total payout ratio').font = BOLD
f = f'=IF($B$3="Bear",B{BEAR_PAYOUT},IF($B$3="Base",B{BASE_PAYOUT},B{BULL_PAYOUT}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_BS = r0
ws.cell(row=r0, column=1, value='Balance sheet / RWA growth (equity retention driver)').font = BOLD
f = f'=IF($B$3="Bear",B{BEAR_BS},IF($B$3="Base",B{BASE_BS},B{BULL_BS}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_TG = r0
ws.cell(row=r0, column=1, value='Terminal growth rate').font = BOLD
f = f'=IF($B$3="Bear",B{BEAR_TG},IF($B$3="Base",B{BASE_TG},B{BULL_TG}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT2; c.font = BLACK; c.border = BORDER
r0 += 1
ACT_KE = r0
ws.cell(row=r0, column=1, value='Cost of equity (CAPM)').font = BOLD
f = f'=IF($B$3="Bear",B{BEAR_KE},IF($B$3="Base",B{BASE_KE},B{BULL_KE}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT2; c.font = BLACK; c.border = BORDER
r0 += 2

ws.cell(row=r0, column=1, value='Preferred dividends, held flat ($mm/yr)').font = BOLD
c = ws.cell(row=r0, column=2, value=PREFERRED_DIV_PROJECTED); c.font = BLUE; c.number_format = USD0; c.border = BORDER
ACT_PREF = r0
r0 += 1
ws.cell(row=r0, column=1, value='Current Price / Book multiple (buyback execution price proxy)').font = BOLD
c = ws.cell(row=r0, column=2, value=GS_PRICE / GS_HISTORY[2025]['bvps']); c.font = BLUE; c.number_format = '0.00x'; c.border = BORDER
ACT_PB = r0
r0 += 1
ws.cell(row=r0, column=1, value='Dividend share of total payout (vs. buybacks), historical avg').font = BOLD
c = ws.cell(row=r0, column=2, value=0.28); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
ACT_DIVSHARE = r0

r0 += 2
methodology = [
"METHODOLOGY NOTE: Goldman Sachs is a bank holding company; a traditional industrial 3-statement",
"model (full working-capital-driven balance sheet) is not the way banks are modeled in practice.",
"Assets/liabilities are financial inventory, not operating capital, so this model uses the",
"industry-standard bank equivalent: an integrated Income Statement -> Common Equity Roll-Forward",
"-> Shares Outstanding walk, all linked through net earnings, payout policy, and required capital",
"retention to support balance-sheet/RWA growth. This is the same architecture used in bank equity",
"research models (see e.g. Damodaran's bank DCF framework) and is fully sufficient to drive EPS,",
"BVPS, ROE, dividends, buybacks and a defensible DCF / reverse-DCF for a G-SIB like GS.",
]
for line in methodology:
    ws.cell(row=r0, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=8)
    r0 += 1

wb.save('/home/claude/gs-valuation/models/02_Three_Statement_Model.xlsx')
print("Assumptions sheet complete.")
print(f"ACT_GROWTH={ACT_GROWTH} ACT_EFF={ACT_EFF} ACT_TAX={ACT_TAX} ACT_PAYOUT={ACT_PAYOUT} "
      f"ACT_BS={ACT_BS} ACT_TG={ACT_TG} ACT_KE={ACT_KE} ACT_PREF={ACT_PREF} ACT_PB={ACT_PB} ACT_DIVSHARE={ACT_DIVSHARE}")
