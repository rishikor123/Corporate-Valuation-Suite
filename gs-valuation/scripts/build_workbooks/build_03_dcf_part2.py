import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=12, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM1 = '#,##0.0'
NUM3 = '0.000'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
years = PROJECTION_YEARS
n = len(years)

ACT_GROWTH, ACT_EFF, ACT_TAX, ACT_PAYOUT, ACT_BS, ACT_TG, ACT_KE = 45, 46, 47, 48, 49, 50, 51
ACT_PREF, MKT_PRICE, MKT_SHARES, MKT_CAP, BASE_EQUITY_2025, BASE_REV_2025 = 53, 56, 57, 58, 59, 60

# ============================= FCFE BUILD ===================================
ws = wb.create_sheet('FCFE Build')
title_block(ws, 'Free Cash Flow to Equity Build (Bank FCFE method), FY2026E-FY2030E', span=7)
ws.column_dimensions['A'].width = 46
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 13
ws.cell(row=3, column=1, value='($ in millions; active scenario per Assumptions!B3)').font = NOTE
ws.cell(row=4, column=1, value='Fiscal Year')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}E')
style_header_row(ws, 4, n + 1)

labels = ['Net revenues', 'Operating expenses', 'Pre-tax earnings', 'Provision for taxes',
          'Net earnings', 'Less: preferred dividends', 'Net earnings applicable to common (NIC)',
          'Beginning common equity', 'Required equity retention (BS/RWA growth x Beg. equity)',
          'Free Cash Flow to Equity (FCFE = NIC - Required retention)',
          'Ending common equity (Beg. equity + Required retention)',
          'Memo: implied retention ratio (Required retention / NIC)']
rows = {}
r = 5
for lbl in labels:
    ws.cell(row=r, column=1, value=lbl).font = BOLD if lbl.startswith('Free Cash') or lbl.startswith('Net earnings applicable') else BLACK
    rows[lbl] = r
    r += 1

R_REV, R_OPEX, R_PRETAX, R_TAX, R_NE, R_PREF, R_NIC, R_BEGEQ, R_RETAIN, R_FCFE, R_ENDEQ, R_RRATIO = \
    [rows[l] for l in labels]

for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    prev = get_column_letter(1 + i)
    acol = get_column_letter(2 + i)
    if i == 0:
        rev_f = f'=Assumptions!$B${BASE_REV_2025}*(1+Assumptions!{acol}{ACT_GROWTH})'
        begeq_f = f'=Assumptions!$B${BASE_EQUITY_2025}'
    else:
        rev_f = f'={prev}{R_REV}*(1+Assumptions!{acol}{ACT_GROWTH})'
        begeq_f = f'={prev}{R_ENDEQ}'
    ws.cell(row=R_REV, column=2 + i, value=rev_f)
    ws.cell(row=R_OPEX, column=2 + i, value=f'={col}{R_REV}*Assumptions!{acol}{ACT_EFF}')
    ws.cell(row=R_PRETAX, column=2 + i, value=f'={col}{R_REV}-{col}{R_OPEX}')
    ws.cell(row=R_TAX, column=2 + i, value=f'={col}{R_PRETAX}*Assumptions!$B${ACT_TAX}')
    ws.cell(row=R_NE, column=2 + i, value=f'={col}{R_PRETAX}-{col}{R_TAX}')
    ws.cell(row=R_PREF, column=2 + i, value=f'=Assumptions!$B${ACT_PREF}')
    ws.cell(row=R_NIC, column=2 + i, value=f'={col}{R_NE}-{col}{R_PREF}')
    ws.cell(row=R_BEGEQ, column=2 + i, value=begeq_f)
    ws.cell(row=R_RETAIN, column=2 + i, value=f'={col}{R_BEGEQ}*Assumptions!$B${ACT_BS}')
    ws.cell(row=R_FCFE, column=2 + i, value=f'={col}{R_NIC}-{col}{R_RETAIN}')
    ws.cell(row=R_ENDEQ, column=2 + i, value=f'={col}{R_BEGEQ}+{col}{R_RETAIN}')
    ws.cell(row=R_RRATIO, column=2 + i, value=f'={col}{R_RETAIN}/{col}{R_NIC}')
    fmt_map = {R_REV: USD0, R_OPEX: USD0, R_PRETAX: USD0, R_TAX: USD0, R_NE: USD0, R_PREF: USD0,
               R_NIC: USD0, R_BEGEQ: USD0, R_RETAIN: USD0, R_FCFE: USD0, R_ENDEQ: USD0, R_RRATIO: PCT1}
    for rr in fmt_map:
        cell = ws.cell(row=rr, column=2 + i)
        cell.number_format = fmt_map[rr]
        cell.border = BORDER
        cell.font = BLACK

r += 2
for line in [
    "Methodology (bank FCFE, per Damodaran's financial-services DCF framework): a bank cannot reinvest",
    "cash flow the way an industrial firm reinvests in capex/working capital; instead, capital is 'used'",
    "by growing risk-weighted assets, which requires retaining common equity to hold leverage/CET1 ratios",
    "roughly constant. FCFE = Net income to common - (required capital retention to fund that growth) =",
    "the maximum a bank can safely return to shareholders via dividends + buybacks without eroding capital.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("FCFE Build sheet complete.")
print(dict(R_REV=R_REV, R_NIC=R_NIC, R_BEGEQ=R_BEGEQ, R_FCFE=R_FCFE, R_ENDEQ=R_ENDEQ))
