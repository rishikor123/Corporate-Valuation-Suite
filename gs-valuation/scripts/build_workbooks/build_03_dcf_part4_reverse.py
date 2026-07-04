import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=13, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM1 = '#,##0.0'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')

ACT_KE = 51
MKT_PRICE, MKT_SHARES, MKT_CAP = 56, 57, 58
R_FCFE = 6  # DCF Valuation sheet FCFE link row

ws = wb.create_sheet('Reverse DCF')
title_block(ws, 'Reverse DCF — What Growth Does the Current Price Already Assume?', span=7)
ws.column_dimensions['A'].width = 48
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 13

ws.cell(row=3, column=1, value='Single-stage (Gordon growth) closed-form solve').font = SUBHDR
ws.cell(row=3, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

r = 4
ws.cell(row=r, column=1, value='Implied current market equity value ($mm)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=Assumptions!$B${MKT_CAP}'); c.number_format = USD0; c.border = BORDER
R_MKTVAL = r
r += 1
ws.cell(row=r, column=1, value='Next-year FCFE, 2026E ($mm) [linked, base case]').font = BLACK
c = ws.cell(row=r, column=2, value=f"='DCF Valuation'!B{R_FCFE}"); c.number_format = USD0; c.border = BORDER
c.font = GREEN
R_FCFE1 = r
r += 1
ws.cell(row=r, column=1, value='Cost of equity, Ke (active scenario)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=Assumptions!$B${ACT_KE}'); c.number_format = PCT2; c.border = BORDER
R_KE = r
r += 1
ws.cell(row=r, column=1, value='Implied FCFE yield (FCFE1 / Market equity value)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=B{R_FCFE1}/B{R_MKTVAL}'); c.number_format = PCT2; c.border = BORDER
R_YIELD = r
r += 1
ws.cell(row=r, column=1, value='=> Implied perpetual FCFE growth rate priced in, g* = Ke - Yield').font = BOLD
c = ws.cell(row=r, column=2, value=f'=B{R_KE}-B{R_YIELD}')
c.number_format = PCT2; c.font = BOLD; c.fill = RESULT_FILL; c.border = BORDER
R_GSTAR = r
r += 2

for line in [
    "Formula: Single-stage Gordon Growth says Price = FCFE1 / (Ke - g). Solving for g given the current",
    "price is algebra: g* = Ke - FCFE1/Price. This is the market-implied 'what has to be true forever'",
    "growth rate. Compare it to (a) GS's own 5-yr revenue CAGR history and (b) nominal US GDP growth",
    "(~4-5%) as sanity checks -- a g* persistently above long-run GDP growth is a classic red flag that",
    "a stock is priced for perfection.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
r += 1

# ---- Sensitivity: implied g* across a range of Ke ----
ws.cell(row=r, column=1, value='Implied growth rate g*, sensitized across cost-of-equity assumptions').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
hdr_row = r
ws.cell(row=r, column=1, value='Cost of equity (Ke)')
ke_range = [0.085, 0.095, 0.105, 0.115, 0.125, 0.135]
for i, kv in enumerate(ke_range):
    c = ws.cell(row=r, column=2 + i, value=kv); c.number_format = PCT1
style_header_row(ws, r, len(ke_range) + 1)
r += 1
ws.cell(row=r, column=1, value='Implied g* (this scenario FCFE1, this Ke)').font = BLACK
for i, kv in enumerate(ke_range):
    col = get_column_letter(2 + i)
    c = ws.cell(row=r, column=2 + i, value=f'={col}{hdr_row}-$B${R_FCFE1}/$B${R_MKTVAL}')
    c.number_format = PCT2; c.border = BORDER
r += 2

# ---- Translate g* into required ROE x retention identity ----
ws.cell(row=r, column=1, value='Sanity check: translating g* into required ROE (sustainable growth identity)').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 1
ws.cell(row=r, column=1, value='Sustainable growth identity: g = ROE x Retention ratio').font = NOTE
r += 1
ws.cell(row=r, column=1, value='Assumed retention ratio (1 - payout), Base Case').font = BLACK
c = ws.cell(row=r, column=2, value='=1-Assumptions!$B$20'); c.number_format = PCT1; c.border = BORDER
R_RETRATIO = r
r += 1
ws.cell(row=r, column=1, value='=> Required ROE to sustain g* at this retention ratio (g* / retention)').font = BOLD
c = ws.cell(row=r, column=2, value=f'=B{R_GSTAR}/B{R_RETRATIO}')
c.number_format = PCT1; c.font = BOLD; c.fill = RESULT_FILL; c.border = BORDER
r += 1
ws.cell(row=r, column=1, value='For reference: GS FY2025A ROE was 15.0%; FY2021 cyclical-peak ROE was 23.0%').font = NOTE
r += 2

for line in [
    "Reading the result: if the required ROE above is comfortably inside GS's own historical range",
    "(7.5% trough in 2023, 23.0% cyclical peak in 2021, 15.0% most recent), the current price is pricing",
    "in a plausible-but-optimistic continuation of the current up-cycle. If it is ABOVE the 2021 peak,",
    "the market is pricing in a structurally higher through-cycle ROE than GS has ever historically",
    "sustained -- a meaningfully more aggressive assumption than the Bull Case in this model.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("Reverse DCF sheet complete.")
