import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=13, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
GS_ROW_FILL = PatternFill('solid', fgColor='DDEBF7')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM2X = '0.00x;(0.00x);"-"'
NUM1 = '#,##0.0'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = Workbook()
ws = wb.active
ws.title = 'Trading Comps'
title_block(ws, 'Comparable Company Analysis — Large-Cap US Banks / Global Investment Banks', span=10)
ws.column_dimensions['A'].width = 24
for col in 'BCDEFGHIJ':
    ws.column_dimensions[col].width = 12

ws.cell(row=3, column=1, value=f'Market data snapshot as of {AS_OF_DATE}; see Data Sources tab for per-company citations').font = NOTE
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)

hdr_row = 5
cols = ['Company', 'Ticker', 'Price', 'Diluted Shares (mm)', 'Market Cap ($mm)', 'BVPS ($)',
        'P / B', 'TTM EPS ($)', 'P / E', 'ROE (ttm/most recent)']
for i, h in enumerate(cols):
    ws.cell(row=hdr_row, column=1 + i, value=h)
style_header_row(ws, hdr_row, len(cols))

companies = [
    ('The Goldman Sachs Group, Inc.', 'GS', GS_PRICE, GS_SHARES_CURRENT_MM, None,
     GS_HISTORY[2025]['bvps'], None, 54.72, None, GS_HISTORY[2025]['roe']),
    ('JPMorgan Chase & Co.', 'JPM', PEERS['JPM']['price'], PEERS['JPM']['shares_mm'], None,
     None, None, PEERS['JPM']['ttm_eps'], None, PEERS['JPM']['roe']),
    ('Morgan Stanley', 'MS', PEERS['MS']['price'], PEERS['MS']['shares_mm'], None,
     None, None, PEERS['MS']['ttm_eps'], None, None),
    ('Bank of America Corp.', 'BAC', PEERS['BAC']['price'], PEERS['BAC']['shares_mm'], None,
     None, None, None, None, None),
    ('Citigroup Inc.', 'C', PEERS['C']['price'], PEERS['C']['shares_mm'], None,
     PEERS['C']['bvps'], None, None, None, PEERS['C']['roe']),
]

r = hdr_row + 1
first_data_row = r
gs_row = r
for name, tick, price, shares, mktcap, bvps, pb, eps, pe, roe in companies:
    ws.cell(row=r, column=1, value=name).font = BLACK
    ws.cell(row=r, column=2, value=tick).font = BOLD
    c = ws.cell(row=r, column=3, value=price); c.font = BLUE; c.number_format = USD2
    c = ws.cell(row=r, column=4, value=shares); c.font = BLUE; c.number_format = NUM1
    c = ws.cell(row=r, column=5, value=f'=C{r}*D{r}'); c.number_format = USD0
    if bvps is not None:
        c = ws.cell(row=r, column=6, value=bvps); c.font = BLUE; c.number_format = USD2
        c = ws.cell(row=r, column=7, value=f'=C{r}/F{r}'); c.number_format = NUM2X
    else:
        ws.cell(row=r, column=6, value='n/d').font = NOTE
        ws.cell(row=r, column=7, value='n/d').font = NOTE
    if eps is not None:
        c = ws.cell(row=r, column=8, value=eps); c.font = BLUE; c.number_format = USD2
        c = ws.cell(row=r, column=9, value=f'=C{r}/H{r}'); c.number_format = NUM2X
    else:
        ws.cell(row=r, column=8, value='n/d').font = NOTE
        ws.cell(row=r, column=9, value='n/d').font = NOTE
    if roe is not None:
        c = ws.cell(row=r, column=10, value=roe); c.font = BLUE; c.number_format = PCT1
    else:
        ws.cell(row=r, column=10, value='n/d').font = NOTE
    for cc in range(1, 11):
        ws.cell(row=r, column=cc).border = BORDER
        if tick == 'GS':
            ws.cell(row=r, column=cc).fill = GS_ROW_FILL
    r += 1
last_data_row = r - 1

r += 1
ws.cell(row=r, column=1, value='Peer group stats (excl. GS)').font = BOLD
peer_first = first_data_row + 1  # JPM row
peer_last = last_data_row
stat_row_start = r
for label, func in [('Peer mean', 'AVERAGE'), ('Peer median', 'MEDIAN')]:
    ws.cell(row=r, column=1, value=label).font = BOLD
    for col_letter, colnum in [('G', 7), ('I', 9), ('J', 10)]:
        rng = f'{col_letter}{peer_first}:{col_letter}{peer_last}'
        c = ws.cell(row=r, column=colnum, value=f'={func}({rng})')
        c.number_format = NUM2X if colnum in (7, 9) else PCT1
        c.border = BORDER
        c.font = BOLD
    r += 1
PEER_MEAN_ROW = stat_row_start
PEER_MEDIAN_ROW = stat_row_start + 1

r += 1
for line in [
    "Peer set: JPMorgan Chase, Morgan Stanley, Bank of America, Citigroup -- the standard large-cap US",
    "bank / global-markets peer group GS itself references in its own investor presentations (GS 2024",
    "Annual Report explicitly names MS, JPM, BAC, C as its primary total-shareholder-return peer set).",
    "'n/d' = not directly sourced for this snapshot; see Data Sources tab. P/E and P/B use TTM figures",
    "where available, which will differ modestly from single-fiscal-year (FY2025A) multiples due to",
    "quarter-mix effects -- treat all multiples here as directional, not decimal-precise.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    r += 1

wb.save('/home/claude/gs-valuation/models/04_Comparable_Company_Analysis.xlsx')
print("Trading Comps sheet complete.")
print(dict(gs_row=gs_row, PEER_MEAN_ROW=PEER_MEAN_ROW, PEER_MEDIAN_ROW=PEER_MEDIAN_ROW))
