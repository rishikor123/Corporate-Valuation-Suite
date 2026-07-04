import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=13, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM2X = '0.00x'

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/04_Comparable_Company_Analysis.xlsx')

GS_ROW = 6
PEER_MEAN_ROW = 12
PEER_MEDIAN_ROW = 13

ws = wb.create_sheet('Implied Valuation')
title_block(ws, 'Implied GS Value per Share — Applying Peer Multiples to GS Financials', span=6)
ws.column_dimensions['A'].width = 40
for col in 'BCDEF':
    ws.column_dimensions[col].width = 14

r = 3
ws.cell(row=r, column=1, value='GS Inputs').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
GS_BVPS_ROW = r
ws.cell(row=r, column=1, value='GS Book value per share ($)').font = BLACK
c = ws.cell(row=r, column=2, value=f"='Trading Comps'!F{GS_ROW}"); c.number_format = USD2; c.border = BORDER
c.font = GREEN
r += 1
GS_EPS_ROW = r
ws.cell(row=r, column=1, value='GS TTM EPS ($)').font = BLACK
c = ws.cell(row=r, column=2, value=f"='Trading Comps'!H{GS_ROW}"); c.number_format = USD2; c.border = BORDER
c.font = GREEN
r += 1
GS_PRICE_ROW = r
ws.cell(row=r, column=1, value='GS current share price ($)').font = BLACK
c = ws.cell(row=r, column=2, value=f"='Trading Comps'!C{GS_ROW}"); c.number_format = USD2; c.border = BORDER
c.font = GREEN
r += 2

ws.cell(row=r, column=1, value='Peer Multiples Applied to GS').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=1, value='Method')
ws.cell(row=r, column=2, value='Peer multiple')
ws.cell(row=r, column=3, value='Applied to GS')
ws.cell(row=r, column=4, value='Implied GS Price/Share ($)')
for c in range(1, 5):
    ws.cell(row=r, column=c).font = HDR; ws.cell(row=r, column=c).fill = HDR_FILL
    ws.cell(row=r, column=c).border = BORDER
    ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', wrap_text=True)
r += 1

rows_out = []
methods = [
    ('P/B, peer mean', f"='Trading Comps'!G{PEER_MEAN_ROW}", f'GS BVPS (${{}})', GS_BVPS_ROW),
    ('P/B, peer median', f"='Trading Comps'!G{PEER_MEDIAN_ROW}", f'GS BVPS (${{}})', GS_BVPS_ROW),
    ('P/E, peer mean', f"='Trading Comps'!I{PEER_MEAN_ROW}", f'GS TTM EPS (${{}})', GS_EPS_ROW),
    ('P/E, peer median', f"='Trading Comps'!I{PEER_MEDIAN_ROW}", f'GS TTM EPS (${{}})', GS_EPS_ROW),
]
implied_rows = []
for label, mult_formula, applied_label, base_row in methods:
    ws.cell(row=r, column=1, value=label).font = BLACK
    c = ws.cell(row=r, column=2, value=mult_formula); c.number_format = NUM2X; c.border = BORDER
    c2 = ws.cell(row=r, column=3, value=f'=B{base_row}'); c2.number_format = USD2; c2.border = BORDER
    c3 = ws.cell(row=r, column=4, value=f'=B{r}*C{r}'); c3.number_format = USD2; c3.border = BORDER
    c3.font = BOLD
    implied_rows.append(r)
    r += 1

r += 1
ws.cell(row=r, column=1, value='Blended average of 4 methods above').font = BOLD
lo, hi = implied_rows[0], implied_rows[-1]
c = ws.cell(row=r, column=4, value=f'=AVERAGE(D{lo}:D{hi})')
c.number_format = USD2; c.font = BOLD; c.fill = RESULT_FILL; c.border = BORDER
BLENDED_ROW = r
r += 1
ws.cell(row=r, column=1, value='Range: min - max of 4 methods').font = BLACK
c = ws.cell(row=r, column=4, value=f'=MIN(D{lo}:D{hi})&" - $"&TEXT(MAX(D{lo}:D{hi}),"0.00")')
c.border = BORDER
r += 2

ws.cell(row=r, column=1, value='Cross-Check vs. Current Price and DCF').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
ws.cell(row=r, column=1, value='Comps-implied price (blended) vs. current price').font = BLACK
c = ws.cell(row=r, column=2, value=f'=D{BLENDED_ROW}/B{GS_PRICE_ROW}-1'); c.number_format = PCT1; c.border = BORDER
r += 1
ws.cell(row=r, column=1, value='DCF Base Case implied price/share (from DCF workbook)').font = BLACK
c = ws.cell(row=r, column=2, value=631.24); c.number_format = USD2; c.font = BLUE; c.border = BORDER
r += 1
ws.cell(row=r, column=1, value='  Source').font = NOTE
ws.cell(row=r, column=2, value="'DCF and Reverse DCF' workbook, DCF Valuation tab, Base Case").font = NOTE
r += 2

for line in [
    "Reading across methods: the comps-implied price uses where the MARKET currently prices GS's peer",
    "set (JPM, MS, BAC, C) and asks 'if GS traded at the same multiple as its peers, what would it be",
    "worth?' -- a relative-value lens. The DCF asks an absolute-value question: 'given this cash-flow",
    "path and discount rate, what is GS intrinsically worth?' When the two disagree meaningfully (as",
    "they do here), the gap is usually explained by (a) the market pricing GS's superior ROE/ROTE",
    "profile at a premium multiple to peers, and/or (b) the DCF's discount rate or terminal growth",
    "being conservative relative to what the market is willing to pay. Both readings belong in the memo.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

order = ['Trading Comps', 'Implied Valuation']
wb._sheets = [wb[s] for s in order]
wb.save('/home/claude/gs-valuation/models/04_Comparable_Company_Analysis.xlsx')
print("Implied Valuation sheet complete.")
