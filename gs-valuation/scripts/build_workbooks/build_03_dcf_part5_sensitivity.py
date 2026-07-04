import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
CORNER_FILL = PatternFill('solid', fgColor='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD2 = '$#,##0.00'
PCT1 = '0.0%'
PCT2 = '0.00%'

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')

MKT_PRICE, MKT_SHARES = 56, 57
FCFE_ROW = "'FCFE Build'!$B$14:$F$14"
LAST_FCFE = "'FCFE Build'!$F$14"
PERIOD_ROW = "'DCF Valuation'!$B$5:$F$5"

ws = wb.create_sheet('Sensitivity Tables')
title_block(ws, 'Sensitivity Analysis — Implied Value per Share ($)', span=9)
ws.column_dimensions['A'].width = 30
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 12

ws.cell(row=3, column=1,
        value='2-Way: Implied share price by Cost of Equity (rows) x Terminal Growth Rate (columns)').font = SUBHDR
ws.cell(row=3, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

ke_vals = [0.085, 0.095, 0.105, 0.108, 0.115, 0.125]
g_vals = [0.015, 0.020, 0.025, 0.030, 0.035, 0.040]

r0 = 5
ws.cell(row=r0, column=1, value='Ke \\ g').font = BOLD
ws.cell(row=r0, column=1).fill = CORNER_FILL
ws.cell(row=r0, column=1).font = Font(bold=True, color='FFFFFF')
for j, g in enumerate(g_vals):
    c = ws.cell(row=r0, column=2 + j, value=g)
    c.number_format = PCT1; c.font = HDR; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
grid_top = r0
for i, kev in enumerate(ke_vals):
    rr = r0 + 1 + i
    kc = ws.cell(row=rr, column=1, value=kev)
    kc.number_format = PCT1; kc.font = HDR; kc.fill = HDR_FILL
    kc.alignment = Alignment(horizontal='center')
    for j, gv in enumerate(g_vals):
        col = get_column_letter(2 + j)
        kecell = f'$A{rr}'
        gcell = f'{col}${grid_top}'
        formula = (f'=(SUMPRODUCT({FCFE_ROW}/(1+{kecell})^{PERIOD_ROW})'
                   f'+(({LAST_FCFE}*(1+{gcell}))/({kecell}-{gcell}))/(1+{kecell})^5)'
                   f'/Assumptions!$B${MKT_SHARES}')
        cell = ws.cell(row=rr, column=2 + j, value=formula)
        cell.number_format = USD2
        cell.border = BORDER

grid_bottom = r0 + len(ke_vals)
rule = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50,
                       mid_color='FFEB84', end_type='max', end_color='63BE7B')
ws.conditional_formatting.add(f'B{grid_top+1}:G{grid_bottom}', rule)

r = grid_bottom + 2
ws.cell(row=r, column=1,
        value=f'Current share price for reference: ${GS_PRICE:,.2f} (highlight cells above/below this manually or eyeball vs. green/red shading)').font = NOTE
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 2

# ---- One-way sensitivities ----
ws.cell(row=r, column=1, value='One-Way Sensitivities (Base Case, holding other drivers constant)').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

def oneway_table(title, header_vals, header_fmt, formula_builder):
    global r
    ws.cell(row=r, column=1, value=title).font = BOLD
    r += 1
    for j, v in enumerate(header_vals):
        c = ws.cell(row=r, column=2 + j, value=v); c.number_format = header_fmt
        c.font = HDR; c.fill = HDR_FILL; c.alignment = Alignment(horizontal='center')
    hdr_r = r
    r += 1
    ws.cell(row=r, column=1, value='Implied value / share ($)').font = BLACK
    for j, v in enumerate(header_vals):
        col = get_column_letter(2 + j)
        f = formula_builder(f'{col}{hdr_r}')
        c = ws.cell(row=r, column=2 + j, value=f)
        c.number_format = USD2; c.border = BORDER
    r += 2

# Terminal growth alone (Ke fixed at active/base 10.8%)
oneway_table('Terminal growth rate (Ke held at 10.8% base case)',
             [0.015, 0.020, 0.025, 0.030, 0.035, 0.040], PCT1,
             lambda gc: (f'=(SUMPRODUCT({FCFE_ROW}/(1.108)^{PERIOD_ROW})'
                         f'+(({LAST_FCFE}*(1+{gc}))/(0.108-{gc}))/(1.108)^5)/Assumptions!$B${MKT_SHARES}'))

# Cost of equity alone (g fixed at 3.0% base case)
oneway_table('Cost of equity (terminal growth held at 3.0% base case)',
             [0.085, 0.095, 0.105, 0.115, 0.125, 0.135], PCT1,
             lambda kc: (f'=(SUMPRODUCT({FCFE_ROW}/(1+{kc})^{PERIOD_ROW})'
                         f'+(({LAST_FCFE}*1.03)/({kc}-0.03))/(1+{kc})^5)/Assumptions!$B${MKT_SHARES}'))

for line in [
    "Reading the grid: at the Base Case (Ke=10.8%, g=3.0%), implied value/share sits well below the",
    "current market price -- consistent with the Reverse DCF finding that the market is pricing in a",
    "higher sustainable growth/ROE than this model's Base Case assumes. Use the grid to see how much",
    "Ke would need to compress (or g rise) to justify the current price -- that combination is the real",
    "underwriting question for anyone buying GS at today's level.",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("Sensitivity Tables sheet complete.")
