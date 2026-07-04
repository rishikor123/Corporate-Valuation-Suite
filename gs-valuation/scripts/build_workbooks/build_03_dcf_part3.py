import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=13, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM1 = '#,##0.0'
NUM3 = '0.000'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
years = PROJECTION_YEARS
n = len(years)

ACT_TG, ACT_KE = 50, 51
MKT_PRICE, MKT_SHARES = 56, 57
R_FCFE, R_REV = 14, 5  # from FCFE Build sheet

ws = wb.create_sheet('DCF Valuation')
title_block(ws, 'DCF Valuation — Present Value of FCFE + Terminal Value', span=7)
ws.column_dimensions['A'].width = 46
for col in 'BCDEFG':
    ws.column_dimensions[col].width = 13

ws.cell(row=3, column=1, value='($ in millions, except per share; active scenario per Assumptions!B3)').font = NOTE
ws.cell(row=4, column=1, value='Fiscal Year')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}E')
style_header_row(ws, 4, n + 1)

r = 5
R_PERIOD = r
ws.cell(row=r, column=1, value='Discount period (t)').font = BLACK
for i in range(n):
    c = ws.cell(row=r, column=2 + i, value=i + 1); c.number_format = '0'; c.border = BORDER
r += 1
R_FCFE_LINK = r
ws.cell(row=r, column=1, value='FCFE [linked from FCFE Build]').font = BOLD
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    c = ws.cell(row=r, column=2 + i, value=f"='FCFE Build'!{col}{R_FCFE}")
    c.number_format = USD0; c.border = BORDER; c.font = GREEN
r += 1
R_DF = r
ws.cell(row=r, column=1, value='Discount factor @ active Ke').font = BLACK
for i in range(n):
    col = get_column_letter(2 + i)
    c = ws.cell(row=r, column=2 + i, value=f'=1/(1+Assumptions!$B${ACT_KE})^{col}{R_PERIOD}')
    c.number_format = NUM3; c.border = BORDER
r += 1
R_PV = r
ws.cell(row=r, column=1, value='PV of FCFE').font = BOLD
for i in range(n):
    col = get_column_letter(2 + i)
    c = ws.cell(row=r, column=2 + i, value=f'={col}{R_FCFE_LINK}*{col}{R_DF}')
    c.number_format = USD0; c.border = BORDER
r += 2

ws.cell(row=r, column=1, value='Valuation Summary').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
R_SUMPV = r
lastcol = get_column_letter(1 + n)
ws.cell(row=r, column=1, value='Sum of PV of explicit-period FCFE (2026E-2030E)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=SUM(B{R_PV}:{lastcol}{R_PV})'); c.number_format = USD0; c.border = BORDER
r += 1
R_TVUNDISC = r
ws.cell(row=r, column=1, value='Terminal value at end of 2030E (Gordon growth)').font = BLACK
c = ws.cell(row=r, column=2,
            value=f'={lastcol}{R_FCFE_LINK}*(1+Assumptions!$B${ACT_TG})/(Assumptions!$B${ACT_KE}-Assumptions!$B${ACT_TG})')
c.number_format = USD0; c.border = BORDER
r += 1
R_TVPV = r
ws.cell(row=r, column=1, value='PV of terminal value').font = BLACK
c = ws.cell(row=r, column=2, value=f'=B{R_TVUNDISC}*{lastcol}{R_DF}'); c.number_format = USD0; c.border = BORDER
r += 1
R_TVPCT = r
ws.cell(row=r, column=1, value='  Terminal value as % of total equity value').font = NOTE
r += 1
R_EQVAL = r
ws.cell(row=r, column=1, value='Implied Equity Value ($mm)').font = BOLD
c = ws.cell(row=r, column=2, value=f'=B{R_SUMPV}+B{R_TVPV}'); c.number_format = USD0; c.border = BORDER
c.font = BOLD
ws.cell(row=R_TVPCT, column=2, value=f'=B{R_TVPV}/B{R_EQVAL}').number_format = PCT1
r += 1
R_SHARES_REF = r
ws.cell(row=r, column=1, value='/ Current diluted shares outstanding (mm)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=Assumptions!$B${MKT_SHARES}'); c.number_format = NUM1; c.border = BORDER
r += 1
R_IMPPRICE = r
ws.cell(row=r, column=1, value='Implied Value per Share ($)').font = BIGBOLD
c = ws.cell(row=r, column=2, value=f'=B{R_EQVAL}/B{R_SHARES_REF}')
c.number_format = USD2; c.font = BIGBOLD; c.fill = RESULT_FILL; c.border = BORDER
r += 1
R_CURPRICE = r
ws.cell(row=r, column=1, value='Current share price ($)').font = BLACK
c = ws.cell(row=r, column=2, value=f'=Assumptions!$B${MKT_PRICE}'); c.number_format = USD2; c.border = BORDER
r += 1
R_UPSIDE = r
ws.cell(row=r, column=1, value='Implied upside / (downside) vs. current price').font = BOLD
c = ws.cell(row=r, column=2, value=f'=B{R_IMPPRICE}/B{R_CURPRICE}-1')
c.number_format = PCT1; c.border = BORDER; c.fill = RESULT_FILL
c.font = BOLD

r += 3
for line in [
    "Read this alongside the Reverse DCF tab: this sheet asks 'what is GS worth given my assumptions?'",
    "The Reverse DCF tab asks the opposite and arguably more useful question for a stock trading near",
    "highs: 'what does the CURRENT price already assume, and is that assumption reasonable?'",
]:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("DCF Valuation sheet complete.")
print(dict(R_FCFE_LINK=R_FCFE_LINK, R_EQVAL=R_EQVAL, R_IMPPRICE=R_IMPPRICE, R_UPSIDE=R_UPSIDE,
           R_TVUNDISC=R_TVUNDISC))
