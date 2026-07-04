import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM1 = '#,##0.0'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/02_Three_Statement_Model.xlsx')
years = PROJECTION_YEARS
n = len(years)

ACT_GROWTH, ACT_EFF, ACT_TAX, ACT_PAYOUT = 38, 39, 40, 41
ACT_BS, ACT_TG, ACT_KE, ACT_PREF, ACT_PB, ACT_DIVSHARE = 42, 43, 44, 46, 47, 48

R_REV, R_GROWTH, R_OPEX, R_PRETAX, R_TAX, R_NE, R_PREF, R_COMMON, R_SHARES, R_EPS, R_MARGIN = \
    6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16
C_BEGEQ, C_NIC, C_TOTPAYOUT, C_DIV, C_BUYBACK, C_ENDEQ, C_BEGSH, C_REPPRICE, C_SHREP, C_ENDSH, \
    C_AVGSH, C_BVPS, C_ROE, C_DPS = 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19

# ============================= INCOME STATEMENT PROJECTION ===================================
ws1 = wb.create_sheet('Income Statement Proj.')
title_block(ws1, 'Projected Income Statement, FY2026E-FY2030E ($ in millions, except per share)', span=7)
ws1.column_dimensions['A'].width = 44
ws1.column_dimensions['B'].width = 13
for col in 'CDEFG':
    ws1.column_dimensions[col].width = 13
ws1.cell(row=4, column=1, value='($ in millions except per share; 2025A shown for reference)').font = NOTE
ws1.cell(row=5, column=1, value='Fiscal Year')
ws1.cell(row=5, column=2, value='2025A')
for i, y in enumerate(years):
    ws1.cell(row=5, column=3 + i, value=f'{y}E')
style_header_row(ws1, 5, n + 2)

labels = {R_REV: 'Net revenues', R_GROWTH: '  % growth YoY', R_OPEX: 'Operating expenses',
          R_PRETAX: 'Pre-tax earnings', R_TAX: 'Provision for taxes', R_NE: 'Net earnings',
          R_PREF: 'Less: preferred dividends', R_COMMON: 'Net earnings applicable to common shareholders',
          R_SHARES: 'Diluted weighted-average shares (mm) [linked]', R_EPS: 'Diluted EPS',
          R_MARGIN: 'Net margin'}
bolds = {R_REV, R_NE, R_COMMON, R_EPS}
for rr, lbl in labels.items():
    ws1.cell(row=rr, column=1, value=lbl).font = BOLD if rr in bolds else BLACK

ws1.cell(row=R_REV, column=2, value=GS_HISTORY[2025]['net_revenues']).number_format = USD0
ws1.cell(row=R_REV, column=2).font = BLUE
ws1.cell(row=R_NE, column=2, value=GS_HISTORY[2025]['net_earnings']).number_format = USD0
ws1.cell(row=R_NE, column=2).font = BLUE
ws1.cell(row=R_PREF, column=2, value=-GS_PREFERRED_DIV[2025]).number_format = USD0
ws1.cell(row=R_PREF, column=2).font = BLUE
ws1.cell(row=R_COMMON, column=2, value=f'=B{R_NE}+B{R_PREF}').number_format = USD0
ws1.cell(row=R_SHARES, column=2, value=GS_SHARES_HISTORY[2025]).number_format = NUM1
ws1.cell(row=R_SHARES, column=2).font = BLUE
ws1.cell(row=R_EPS, column=2, value=f'=B{R_COMMON}/B{R_SHARES}').number_format = USD2
ws1.cell(row=R_MARGIN, column=2, value=f'=B{R_NE}/B{R_REV}').number_format = PCT1
for rr in [R_COMMON, R_EPS, R_MARGIN]:
    ws1.cell(row=rr, column=2).font = BLACK

for i, y in enumerate(years):
    col = get_column_letter(3 + i)
    prev = get_column_letter(2 + i)
    acol = get_column_letter(2 + i)
    ws1.cell(row=R_REV, column=3 + i, value=f'={prev}{R_REV}*(1+Assumptions!{acol}{ACT_GROWTH})')
    ws1.cell(row=R_GROWTH, column=3 + i, value=f'=Assumptions!{acol}{ACT_GROWTH}')
    ws1.cell(row=R_OPEX, column=3 + i, value=f'={col}{R_REV}*Assumptions!{acol}{ACT_EFF}')
    ws1.cell(row=R_PRETAX, column=3 + i, value=f'={col}{R_REV}-{col}{R_OPEX}')
    ws1.cell(row=R_TAX, column=3 + i, value=f'={col}{R_PRETAX}*Assumptions!$B${ACT_TAX}')
    ws1.cell(row=R_NE, column=3 + i, value=f'={col}{R_PRETAX}-{col}{R_TAX}')
    ws1.cell(row=R_PREF, column=3 + i, value=f'=-Assumptions!$B${ACT_PREF}')
    ws1.cell(row=R_COMMON, column=3 + i, value=f'={col}{R_NE}+{col}{R_PREF}')
    ws1.cell(row=R_SHARES, column=3 + i, value=f"='Capital & Shares'!{col}{C_AVGSH}")
    ws1.cell(row=R_EPS, column=3 + i, value=f'={col}{R_COMMON}/{col}{R_SHARES}')
    ws1.cell(row=R_MARGIN, column=3 + i, value=f'={col}{R_NE}/{col}{R_REV}')
    fmt_map = {R_REV: USD0, R_GROWTH: PCT1, R_OPEX: USD0, R_PRETAX: USD0, R_TAX: USD0, R_NE: USD0,
               R_PREF: USD0, R_COMMON: USD0, R_SHARES: NUM1, R_EPS: USD2, R_MARGIN: PCT1}
    for rr in fmt_map:
        cell = ws1.cell(row=rr, column=3 + i)
        cell.number_format = fmt_map[rr]
        cell.border = BORDER
        cell.font = GREEN if rr in (R_GROWTH, R_SHARES) else BLACK

r = 18
for line in [
    "Formula flow: Net revenues grow off Assumptions active drivers -> Opex = Revenue x efficiency ratio",
    "-> Pre-tax = Revenue - Opex -> Tax = Pre-tax x tax rate -> Net earnings -> less preferred dividends",
    "-> Net earnings to common. Diluted shares link FROM the Capital & Shares tab (green cells = links).",
    "No circularity: shares/BVPS depend on Net earnings to common, which is computed independently above.",
]:
    ws1.cell(row=r, column=1, value=line).font = NOTE
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

# ============================= CAPITAL & SHARES ROLL-FORWARD ===================================
ws2 = wb.create_sheet('Capital & Shares')
title_block(ws2, 'Common Equity & Diluted Shares Roll-Forward, FY2026E-FY2030E', span=7)
ws2.column_dimensions['A'].width = 44
ws2.column_dimensions['B'].width = 13
for col in 'CDEFG':
    ws2.column_dimensions[col].width = 13
ws2.cell(row=4, column=1, value='($ in millions except per share; shares in millions; 2025A shown for reference)').font = NOTE
ws2.cell(row=5, column=1, value='Fiscal Year')
ws2.cell(row=5, column=2, value='2025A')
for i, y in enumerate(years):
    ws2.cell(row=5, column=3 + i, value=f'{y}E')
style_header_row(ws2, 5, n + 2)

labels2 = {C_BEGEQ: 'Beginning common equity', C_NIC: 'Net earnings applicable to common [linked]',
           C_TOTPAYOUT: 'Total capital returned (payout ratio x NI to common)',
           C_DIV: '  Common dividends', C_BUYBACK: '  Share repurchases ($)',
           C_ENDEQ: 'Ending common equity', C_BEGSH: 'Beginning diluted shares (mm)',
           C_REPPRICE: 'Assumed repurchase price / share', C_SHREP: 'Shares repurchased (mm)',
           C_ENDSH: 'Ending diluted shares (mm)', C_AVGSH: 'Average diluted shares (mm) [feeds EPS]',
           C_BVPS: 'Book value per share (ending)', C_ROE: 'Return on average common equity',
           C_DPS: 'Dividend per share'}
bolds2 = {C_ENDEQ, C_ENDSH, C_BVPS, C_ROE}
for rr, lbl in labels2.items():
    ws2.cell(row=rr, column=1, value=lbl).font = BOLD if rr in bolds2 else BLACK

base_equity_2025 = GS_HISTORY[2025]['bvps'] * GS_SHARES_HISTORY[2025]
ws2.cell(row=C_ENDEQ, column=2, value=round(base_equity_2025, 1)).number_format = USD0
ws2.cell(row=C_ENDEQ, column=2).font = BLUE
ws2.cell(row=C_ENDSH, column=2, value=GS_SHARES_HISTORY[2025]).number_format = NUM1
ws2.cell(row=C_ENDSH, column=2).font = BLUE
ws2.cell(row=C_BVPS, column=2, value=f'=B{C_ENDEQ}/B{C_ENDSH}').number_format = USD2
ws2.cell(row=C_BVPS, column=2).font = BLACK

for i, y in enumerate(years):
    col = get_column_letter(3 + i)
    prev = get_column_letter(2 + i)
    ws2.cell(row=C_BEGEQ, column=3 + i, value=f'={prev}{C_ENDEQ}')
    ws2.cell(row=C_NIC, column=3 + i, value=f"='Income Statement Proj.'!{col}{R_COMMON}")
    ws2.cell(row=C_TOTPAYOUT, column=3 + i, value=f'={col}{C_NIC}*Assumptions!$B${ACT_PAYOUT}')
    ws2.cell(row=C_DIV, column=3 + i, value=f'={col}{C_TOTPAYOUT}*Assumptions!$B${ACT_DIVSHARE}')
    ws2.cell(row=C_BUYBACK, column=3 + i, value=f'={col}{C_TOTPAYOUT}-{col}{C_DIV}')
    ws2.cell(row=C_ENDEQ, column=3 + i, value=f'={col}{C_BEGEQ}+{col}{C_NIC}-{col}{C_TOTPAYOUT}')
    ws2.cell(row=C_BEGSH, column=3 + i, value=f'={prev}{C_ENDSH}')
    ws2.cell(row=C_REPPRICE, column=3 + i, value=f'=({col}{C_BEGEQ}/{col}{C_BEGSH})*Assumptions!$B${ACT_PB}')
    ws2.cell(row=C_SHREP, column=3 + i, value=f'={col}{C_BUYBACK}/{col}{C_REPPRICE}')
    ws2.cell(row=C_ENDSH, column=3 + i, value=f'={col}{C_BEGSH}-{col}{C_SHREP}')
    ws2.cell(row=C_AVGSH, column=3 + i, value=f'=AVERAGE({col}{C_BEGSH},{col}{C_ENDSH})')
    ws2.cell(row=C_BVPS, column=3 + i, value=f'={col}{C_ENDEQ}/{col}{C_ENDSH}')
    ws2.cell(row=C_ROE, column=3 + i, value=f'={col}{C_NIC}/AVERAGE({col}{C_BEGEQ},{col}{C_ENDEQ})')
    ws2.cell(row=C_DPS, column=3 + i, value=f'={col}{C_DIV}/{col}{C_AVGSH}')
    fmt_map2 = {C_BEGEQ: USD0, C_NIC: USD0, C_TOTPAYOUT: USD0, C_DIV: USD0, C_BUYBACK: USD0,
                C_ENDEQ: USD0, C_BEGSH: NUM1, C_REPPRICE: USD2, C_SHREP: NUM1, C_ENDSH: NUM1,
                C_AVGSH: NUM1, C_BVPS: USD2, C_ROE: PCT1, C_DPS: USD2}
    for rr in fmt_map2:
        cell = ws2.cell(row=rr, column=3 + i)
        cell.number_format = fmt_map2[rr]
        cell.border = BORDER
        cell.font = GREEN if rr == C_NIC else BLACK

r = 21
for line in [
    "Equity roll-forward: Ending equity = Beginning equity + Net income to common - Total capital",
    "returned (dividends + buybacks). Buyback $ / assumed repurchase price (BVPS x current P/B multiple)",
    "= shares repurchased -> Ending shares -> BVPS and ROE fall out mechanically. Average shares feed",
    "diluted EPS on the Income Statement Proj. tab (see green-linked cells there).",
]:
    ws2.cell(row=r, column=1, value=line).font = NOTE
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

wb.save('/home/claude/gs-valuation/models/02_Three_Statement_Model.xlsx')
print("IS Proj. and Capital & Shares sheets complete.")
print(dict(R_REV=R_REV, R_COMMON=R_COMMON, R_EPS=R_EPS, C_ENDEQ=C_ENDEQ, C_ENDSH=C_ENDSH,
           C_BVPS=C_BVPS, C_ROE=C_ROE, C_DIV=C_DIV, C_BUYBACK=C_BUYBACK, C_AVGSH=C_AVGSH, C_DPS=C_DPS))
