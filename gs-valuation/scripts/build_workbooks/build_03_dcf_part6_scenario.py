import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLACK = Font(name='Arial', size=9, color='000000')
BOLD = Font(name='Arial', size=10, bold=True)
BIGBOLD = Font(name='Arial', size=13, bold=True, color='1F4E78')
HDR = Font(name='Arial', size=9, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
RESULT_FILL = PatternFill('solid', fgColor='FFF2CC')
SCEN_FILL = {'Bear': PatternFill('solid', fgColor='F8CBAD'),
             'Base': PatternFill('solid', fgColor='D9E1F2'),
             'Bull': PatternFill('solid', fgColor='C6E0B4')}
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'

def title_block(ws, text, span=17):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
years = PROJECTION_YEARS
n = len(years)

BLOCK_ROWS = {'Bear': dict(growth=7, eff=8, tax=9, payout=10, bs=11, tg=12, ke=13),
              'Base': dict(growth=17, eff=18, tax=19, payout=20, bs=21, tg=22, ke=23),
              'Bull': dict(growth=27, eff=28, tax=29, payout=30, bs=31, tg=32, ke=33)}
ACT_PREF, MKT_PRICE, MKT_SHARES, BASE_EQUITY_2025, BASE_REV_2025 = 53, 56, 57, 59, 60

ws = wb.create_sheet('Scenario Valuation')
title_block(ws, 'Bull / Base / Bear — Independent DCF Valuation Comparison', span=17)
ws.column_dimensions['A'].width = 32
for col_i in range(2, 18):
    ws.column_dimensions[get_column_letter(col_i)].width = 11

row_scen = 3
row_year = 4
col = 2
scen_start_col = {}
for scen in ['Bear', 'Base', 'Bull']:
    scen_start_col[scen] = col
    ws.merge_cells(start_row=row_scen, start_column=col, end_row=row_scen, end_column=col + n - 1)
    c = ws.cell(row=row_scen, column=col, value=f'{scen} Case')
    c.font = HDR; c.fill = SCEN_FILL[scen]; c.alignment = Alignment(horizontal='center')
    for i, y in enumerate(years):
        yc = ws.cell(row=row_year, column=col + i, value=f'{y}E')
        yc.font = HDR; yc.fill = HDR_FILL; yc.alignment = Alignment(horizontal='center'); yc.border = BORDER
    col += n + 1

metrics = ['Net revenues ($mm)', 'Net earnings to common ($mm)', 'Beginning common equity ($mm)',
           'Required equity retention ($mm)', 'FCFE ($mm)', 'Discount factor', 'PV of FCFE ($mm)']
r0 = 6
metric_row = {}
for m in metrics:
    ws.cell(row=r0, column=1, value=m).font = BOLD if 'FCFE' in m and 'PV' not in m else BLACK
    metric_row[m] = r0
    r0 += 1

for scen in ['Bear', 'Base', 'Bull']:
    sc = scen_start_col[scen]
    br = BLOCK_ROWS[scen]
    for i, y in enumerate(years):
        col_letter = get_column_letter(sc + i)
        acol = get_column_letter(2 + i)
        prev_letter = get_column_letter(sc + i - 1) if i > 0 else None

        if i == 0:
            rev_f = f'=Assumptions!$B${BASE_REV_2025}*(1+Assumptions!{acol}{br["growth"]})'
            begeq_f = f'=Assumptions!$B${BASE_EQUITY_2025}'
        else:
            rev_f = f'={prev_letter}{metric_row["Net revenues ($mm)"]}*(1+Assumptions!{acol}{br["growth"]})'
            begeq_f = f'={prev_letter}{metric_row["Beginning common equity ($mm)"]}*(1+Assumptions!$B${br["bs"]})'
        rc = ws.cell(row=metric_row['Net revenues ($mm)'], column=sc + i, value=rev_f)
        rc.number_format = USD0; rc.border = BORDER; rc.font = BLACK

        rev_ref = f'{col_letter}{metric_row["Net revenues ($mm)"]}'
        nic_formula = (f'=({rev_ref}*(1-Assumptions!{acol}{br["eff"]}))*(1-Assumptions!$B${br["tax"]})'
                       f'-Assumptions!$B${ACT_PREF}')
        nc = ws.cell(row=metric_row['Net earnings to common ($mm)'], column=sc + i, value=nic_formula)
        nc.number_format = USD0; nc.border = BORDER; nc.font = BLACK

        beq = ws.cell(row=metric_row['Beginning common equity ($mm)'], column=sc + i, value=begeq_f)
        beq.number_format = USD0; beq.border = BORDER; beq.font = BLACK

        beq_ref = f'{col_letter}{metric_row["Beginning common equity ($mm)"]}'
        retain_formula = f'={beq_ref}*Assumptions!$B${br["bs"]}'
        rt = ws.cell(row=metric_row['Required equity retention ($mm)'], column=sc + i, value=retain_formula)
        rt.number_format = USD0; rt.border = BORDER; rt.font = BLACK

        nic_ref = f'{col_letter}{metric_row["Net earnings to common ($mm)"]}'
        retain_ref = f'{col_letter}{metric_row["Required equity retention ($mm)"]}'
        fcfe_formula = f'={nic_ref}-{retain_ref}'
        fc = ws.cell(row=metric_row['FCFE ($mm)'], column=sc + i, value=fcfe_formula)
        fc.number_format = USD0; fc.border = BORDER; fc.font = BOLD

        df_formula = f'=1/(1+Assumptions!$B${br["ke"]})^{i+1}'
        dfc = ws.cell(row=metric_row['Discount factor'], column=sc + i, value=df_formula)
        dfc.number_format = '0.000'; dfc.border = BORDER; dfc.font = BLACK

        fcfe_ref = f'{col_letter}{metric_row["FCFE ($mm)"]}'
        df_ref = f'{col_letter}{metric_row["Discount factor"]}'
        pv_formula = f'={fcfe_ref}*{df_ref}'
        pvc = ws.cell(row=metric_row['PV of FCFE ($mm)'], column=sc + i, value=pv_formula)
        pvc.number_format = USD0; pvc.border = BORDER; pvc.font = BLACK

r0 += 2
ws.cell(row=r0, column=1, value='Valuation Output').font = SUBHDR
ws.cell(row=r0, column=1).fill = SUBHDR_FILL
r0 += 1
ws.cell(row=r0, column=1, value='Scenario').font = HDR
ws.cell(row=r0, column=1).fill = HDR_FILL
out_cols = ['Sum PV explicit FCFE ($mm)', 'Terminal value, undiscounted ($mm)', 'PV of terminal value ($mm)',
            'Implied Equity Value ($mm)', 'Implied Value / Share ($)', 'vs. Current Price']
for j, m in enumerate(out_cols):
    c = ws.cell(row=r0, column=2 + j, value=m); c.font = HDR; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
r0 += 1
for scen in ['Bear', 'Base', 'Bull']:
    sc = scen_start_col[scen]
    first_col = get_column_letter(sc)
    last_col = get_column_letter(sc + n - 1)
    br = BLOCK_ROWS[scen]
    ws.cell(row=r0, column=1, value=scen).font = BOLD
    ws.cell(row=r0, column=1).fill = SCEN_FILL[scen]

    sumpv_f = f'=SUM({first_col}{metric_row["PV of FCFE ($mm)"]}:{last_col}{metric_row["PV of FCFE ($mm)"]})'
    c1 = ws.cell(row=r0, column=2, value=sumpv_f); c1.number_format = USD0; c1.border = BORDER

    tv_f = (f'={last_col}{metric_row["FCFE ($mm)"]}*(1+Assumptions!$B${br["tg"]})'
            f'/(Assumptions!$B${br["ke"]}-Assumptions!$B${br["tg"]})')
    c2 = ws.cell(row=r0, column=3, value=tv_f); c2.number_format = USD0; c2.border = BORDER

    tvpv_f = f'=C{r0}*{last_col}{metric_row["Discount factor"]}'
    c3 = ws.cell(row=r0, column=4, value=tvpv_f); c3.number_format = USD0; c3.border = BORDER

    eqval_f = f'=B{r0}+D{r0}'
    c4 = ws.cell(row=r0, column=5, value=eqval_f); c4.number_format = USD0; c4.border = BORDER
    c4.font = BOLD

    price_f = f'=E{r0}/Assumptions!$B${MKT_SHARES}'
    c5 = ws.cell(row=r0, column=6, value=price_f); c5.number_format = USD2; c5.border = BORDER
    c5.font = BOLD; c5.fill = RESULT_FILL

    vs_f = f'=F{r0}/Assumptions!$B${MKT_PRICE}-1'
    c6 = ws.cell(row=r0, column=7, value=vs_f); c6.number_format = PCT1; c6.border = BORDER
    r0 += 1

r0 += 2
for line in [
    "Each scenario uses its own revenue growth, efficiency ratio, tax rate, balance-sheet growth,",
    "terminal growth AND cost of equity (wider discount rate in Bear to reflect higher perceived risk,",
    "tighter in Bull) -- consistent with how sell-side scenario analysis typically treats risk, not just",
    "growth. This is a fully independent recompute per scenario, not a simple +/- shock to the Base Case.",
]:
    ws.cell(row=r0, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=10)
    r0 += 1

order = ['Assumptions', 'FCFE Build', 'DCF Valuation', 'Reverse DCF', 'Sensitivity Tables', 'Scenario Valuation']
wb._sheets = [wb[s] for s in order]
wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("Scenario Valuation sheet complete.")
