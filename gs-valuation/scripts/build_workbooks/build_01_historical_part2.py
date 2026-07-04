import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=7):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']
    c.value = text
    c.font = TITLE
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/01_Historical_Financial_Analysis.xlsx')
years = [2021, 2022, 2023, 2024, 2025]

# ============================= KEY RATIOS ===================================
ws = wb.create_sheet('Key Ratios')
title_block(ws, 'Key Profitability & Efficiency Ratios, FY2021-FY2025', span=6)
ws.column_dimensions['A'].width = 36
for col in 'BCDEF':
    ws.column_dimensions[col].width = 13

ws.cell(row=4, column=1, value='Metric')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}A')
style_header_row(ws, 4, 6)

r = 5
def hard_row(label, values, fmt, bold=False):
    global r
    ws.cell(row=r, column=1, value=label).font = BOLD if bold else BLACK
    for i, v in enumerate(values):
        cell = ws.cell(row=r, column=2 + i, value=v)
        cell.number_format = fmt
        cell.font = NOTE if v in ('n/d', 'n/a') else BLUE
        cell.border = BORDER
    r += 1
    return r - 1

row_roe = hard_row('Return on average common equity (ROE)', [GS_HISTORY[y]['roe'] for y in years], PCT1, bold=True)
row_rote = hard_row('Return on average tangible common equity (ROTE)', [GS_HISTORY[y]['rote'] for y in years], PCT1)
row_eff = hard_row('Efficiency ratio (Opex / Net revenues)',
                    [GS_HISTORY[y]['efficiency_ratio'] if GS_HISTORY[y]['efficiency_ratio'] else 'n/d' for y in years], PCT1)
row_bvps = hard_row('Book value per share', [GS_HISTORY[y]['bvps'] for y in years], USD2, bold=True)
r += 1

ws.cell(row=r, column=1, value='  BVPS growth YoY').font = BLACK
for i in range(1, len(years)):
    col = get_column_letter(2 + i); prev = get_column_letter(1 + i)
    c = ws.cell(row=r, column=2 + i, value=f'={col}{row_bvps}/{prev}{row_bvps}-1')
    c.number_format = PCT1; c.font = BLACK; c.border = BORDER
ws.cell(row=r, column=2, value='n/a').font = NOTE
r += 2

ws.cell(row=r, column=1, value='ROE decomposition (simplified DuPont)').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
r += 1
ws.cell(row=r, column=1, value='Net margin (Net earnings / Net revenues)').font = BLACK
for i, y in enumerate(years):
    col = get_column_letter(2 + i)
    c = ws.cell(row=r, column=2 + i, value=f"='Income Statement'!{col}{15}")
    c.number_format = PCT1; c.font = GREEN; c.border = BORDER
r += 1
ws.cell(row=r, column=1, value='Note: full DuPont (asset turnover x leverage) requires average total').font = NOTE
r += 1
ws.cell(row=r, column=1, value='assets/equity detail not broken out in earnings-release summaries; see 10-K').font = NOTE
r += 1
ws.cell(row=r, column=1, value='Note 20 (Capital) for full regulatory-capital ratio detail if extending this model.').font = NOTE

# Bar-style commentary rows
r += 2
ws.cell(row=r, column=1, value='Reading the trend').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
commentary = [
 "2021 was a cyclical peak (record capital-markets activity, ROE 23.0%) — not a representative run-rate.",
 "2022-2023 show the down-cycle: capital markets activity slowed and the firm absorbed costs tied to the",
 "unwound consumer strategy (GreenSky, GM Card, Marcus), pushing ROE to 7.5% and efficiency ratio to 74.6% in 2023.",
 "2024-2025 show a clean re-acceleration: ROE recovered to 12.7% then 15.0%, roughly back to the firm's",
 "through-cycle 14-16% target band, on Global Banking & Markets strength and a narrower consumer footprint.",
]
for line in commentary:
    ws.cell(row=r, column=1, value=line).font = Font(name='Arial', size=9)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

# ============================= PER SHARE & CAPITAL ===================================
ws = wb.create_sheet('Per Share & Capital')
title_block(ws, 'Per-Share Metrics, Capital Return & Share Count, FY2021-FY2025', span=6)
ws.column_dimensions['A'].width = 36
for col in 'BCDEF':
    ws.column_dimensions[col].width = 13

ws.cell(row=4, column=1, value='Metric')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}A')
style_header_row(ws, 4, 6)
r = 5
row_shares = hard_row2 = None

def hard_row2(label, values, fmt, bold=False, green_from_is=False):
    global r
    ws.cell(row=r, column=1, value=label).font = BOLD if bold else BLACK
    for i, v in enumerate(values):
        cell = ws.cell(row=r, column=2 + i, value=v)
        cell.number_format = fmt
        cell.font = BLUE
        cell.border = BORDER
    r += 1
    return r - 1

row_eps = hard_row2('Diluted EPS', [GS_HISTORY[y]['diluted_eps'] for y in years], USD2, bold=True)
row_bvps2 = hard_row2('Book value per share', [GS_HISTORY[y]['bvps'] for y in years], USD2, bold=True)
row_shr = hard_row2('Diluted weighted-avg shares outstanding (mm)', [GS_SHARES_HISTORY[y] for y in years], '#,##0.0')
r += 1
ws.cell(row=r, column=1, value='  Shares outstanding change YoY').font = BLACK
for i in range(1, len(years)):
    col = get_column_letter(2 + i); prev = get_column_letter(1 + i)
    c = ws.cell(row=r, column=2 + i, value=f'={col}{row_shr}/{prev}{row_shr}-1')
    c.number_format = PCT1; c.font = BLACK; c.border = BORDER
ws.cell(row=r, column=2, value='n/a').font = NOTE
r += 2

ws.cell(row=r, column=1, value='Capital return context').font = SUBHDR
ws.cell(row=r, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
notes = [
    ('FY2024 capital returned to shareholders ($bn)', 11.80, '$#,##0.00'),
    ('  of which share repurchases ($bn)', 8.00, '$#,##0.00'),
    ('  of which common dividends ($bn)', 3.80, '$#,##0.00'),
    ('FY2025 4Q common shares repurchased (mm)', 18.9, '#,##0.0'),
    ('FY2025 4Q repurchase cost ($bn)', 12.36, '$#,##0.00'),
]
for label, val, fmt in notes:
    ws.cell(row=r, column=1, value=label).font = BLACK
    c = ws.cell(row=r, column=2, value=val); c.font = BLUE; c.number_format = fmt
    r += 1

r += 1
ws.cell(row=r, column=1, value='Source: GS FY2024 Earnings Release (Jan 15 2025); GS 4Q25 Earnings').font = NOTE
r += 1
ws.cell(row=r, column=1, value='Results Presentation (Jan 15 2026). Diluted share count fell from 355.0mm').font = NOTE
r += 1
ws.cell(row=r, column=1, value='(2021) to 300.6mm (2025), a -15.3% reduction over 4 years, almost entirely').font = NOTE
r += 1
ws.cell(row=r, column=1, value='via buybacks — this buyback cadence is a key DCF/reverse-DCF driver.').font = NOTE

# ============================= SEGMENT ANALYSIS ===================================
ws = wb.create_sheet('Segment Analysis')
title_block(ws, 'Net Revenues by Business Segment ($ in millions)', span=6)
ws.column_dimensions['A'].width = 34
for col in 'BCDEF':
    ws.column_dimensions[col].width = 13
ws.cell(row=4, column=1, value='Segment')
for i, y in enumerate(years):
    ws.cell(row=4, column=2 + i, value=f'{y}A')
style_header_row(ws, 4, 6)
r = 5
# Note: GS restructured segments over the period (3-segment -> 4-segment -> 3-segment in 4Q25).
# Values below use the segment structure as originally reported each year (not restated), consistent
# with as-filed earnings releases. See note below the table.
seg_data = [
    ('Investment Banking (2021 structure)', 14876, None, None, None, None),
    ('Global Markets (2021 structure)', 22077, None, None, None, None),
    ('Asset Management (2021 structure)', 14916, None, None, None, None),
    ('Consumer & Wealth Management (2021 structure)', 7470, None, None, None, None),
    ('Global Banking & Markets (2022-25 structure)', None, 32487, None, None, 41500),
    ('Asset & Wealth Management (2022-25 structure)', None, 13376, None, None, None),
    ('Platform Solutions (2022-25 structure)', None, 1502, None, None, None),
    ('Total net revenues (as reported)', 59339, 47365, 46254, 53510, 58280),
]
for label, *vals in seg_data:
    ws.cell(row=r, column=1, value=label).font = BOLD if 'Total' in label else BLACK
    for i, v in enumerate(vals):
        cell = ws.cell(row=r, column=2 + i, value=v if v is not None else 'n/d')
        cell.number_format = USD0
        cell.font = NOTE if v is None else (BLUE if 'Total' not in label else BOLD)
        cell.border = BORDER
    r += 1

r += 1
seg_note_lines = [
    "Note: Goldman Sachs changed its business-segment structure twice in this window: 4-segment",
    "(Investment Banking / Global Markets / Asset Management / Consumer & Wealth Mgmt) through FY2021,",
    "reorganized to 3 segments (Global Banking & Markets / Asset & Wealth Mgmt / Platform Solutions) for",
    "FY2022-FY2025, then to a further-simplified structure starting 4Q25 (see Form 8-K filed Jan 8, 2026).",
    "Prior-period segment figures are NOT restated here to the latest structure — pull the FY2025 10-K",
    "Note 25 (Business Segments) 5-year recast table before using this tab for segment-level modeling.",
    "The only fully reliable cross-year comparison is Total net revenues (as reported), used throughout",
    "the rest of this workbook and the 3-statement model.",
]
for line in seg_note_lines:
    ws.cell(row=r, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

# ============================= DATA SOURCES ===================================
ws = wb.create_sheet('Data Sources')
title_block(ws, 'Data Sources & Citations', span=4)
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 70
r = 3
ws.cell(row=r, column=1, value='Item').font = BOLD
ws.cell(row=r, column=2, value='Source').font = BOLD
r += 1
srcs = [
    ('FY2021-FY2025 net revenues, net earnings, diluted EPS, ROE, ROTE', 'GS Full Year & 4Q Earnings Results press releases (Form 8-K, Ex-99.1/99.2), SEC EDGAR, respective January each year'),
    ('FY2023-FY2024 efficiency ratio, operating expenses', 'GS FY2024 10-K analysis summary; GS 2024 Annual Report'),
    ('FY2021-FY2025 book value per share', 'GS Full Year & 4Q Earnings Results press releases; FY2024 BVPS derived from disclosed +6.2% YoY growth to $357.60 stated in the 4Q25 release'),
    ('Diluted weighted-average shares outstanding', 'Estimated from Net earnings applicable to common / reported diluted EPS each year; cross-checked against 313.9mm shares outstanding disclosed in 3Q24 10-Q (Oct 18, 2024)'),
    ('Preferred dividends', 'Estimated at ~4% of net earnings, consistent with GS preferred stock program disclosures; verify exact figure in each 10-K Note 21 (Earnings Per Common Share) before use in a live deliverable'),
    ('Current share price, 52-week range', 'Robinhood GS quote, 2026-07-02'),
    ('Current shares outstanding, beta, dividend', 'Google Finance GS:NYSE quote, snapshot ~2026-07-02'),
    ('10-Year US Treasury yield', 'Trading Economics / MacroMicro, 2026-07-02 (4.48%)'),
    ('Capital return context (FY2024 buybacks/dividends)', 'GS FY2024 Full Year Earnings Release, Jan 15 2025'),
    ('4Q25 buyback detail (18.9mm shares, $12.36bn)', 'GS 4Q25 Earnings Results Presentation, Jan 15 2026'),
    ('Total assets (~$1.809T)', 'companiesmarketcap.com, GS total assets snapshot, 2026'),
    ('Peer comps (JPM, MS, BAC, C)', 'See "Comparable Company Analysis" workbook, Data Sources tab'),
]
for item, src in srcs:
    ws.cell(row=r, column=1, value=item).font = BLACK
    c = ws.cell(row=r, column=2, value=src); c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 30
    r += 1

r += 1
disclaim = ("DISCLAIMER: This workbook is a portfolio / educational project, not investment research. "
            "All figures are sourced from public filings and secondary market-data sites as of the dates "
            "noted above and may contain transcription or staleness error. Verify all inputs against the "
            "primary source (SEC EDGAR filing) before relying on this model for any real decision.")
ws.cell(row=r, column=1, value=disclaim).font = NOTE
ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=6)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical='top')

# reorder sheets
order = ['Cover', 'Income Statement', 'Key Ratios', 'Per Share & Capital', 'Segment Analysis', 'Data Sources']
wb._sheets = [wb[s] for s in order]

wb.save('/home/claude/gs-valuation/models/01_Historical_Financial_Analysis.xlsx')
print("Workbook 1 complete with all sheets.")
