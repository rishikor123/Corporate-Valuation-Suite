import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=10, color='000000')
GREEN = Font(name='Arial', size=10, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
BOLDBLUE = Font(name='Arial', size=11, bold=True, color='0000FF')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SUBHDR_FILL = PatternFill('solid', fgColor='D9E1F2')
SUBHDR = Font(name='Arial', size=10, bold=True, color='1F4E78')
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
YELLOW = PatternFill('solid', fgColor='FFFF00')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM1 = '#,##0.0'
NUM2X = '0.00x'

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def title_block(ws, text, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = Workbook()
years = PROJECTION_YEARS
n = len(years)

# ============================= ASSUMPTIONS (self-contained copy) ===================================
ws = wb.active
ws.title = 'Assumptions'
title_block(ws, 'DCF Model Assumptions — Bear / Base / Bull Drivers (self-contained)', span=8)
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 14
for col in 'CDEFG':
    ws.column_dimensions[col].width = 11

ws['A3'] = 'ACTIVE SCENARIO (type Bear / Base / Bull):'
ws['A3'].font = BOLD
ws['B3'] = 'Base'
ws['B3'].font = BOLDBLUE
ws['B3'].fill = YELLOW
ws['B3'].border = BORDER
dv = DataValidation(type='list', formula1='"Bear,Base,Bull"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws['B3'])

r0 = 5
BLOCK_ROWS = {}
for scen in ['Bear', 'Base', 'Bull']:
    ws.cell(row=r0, column=1, value=f'{scen} Case drivers').font = SUBHDR
    ws.cell(row=r0, column=1).fill = SUBHDR_FILL
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=1 + n)
    r0 += 1
    hdr_row = r0
    ws.cell(row=r0, column=1, value='Fiscal Year')
    for i, y in enumerate(years):
        ws.cell(row=r0, column=2 + i, value=f'{y}E')
    style_header_row(ws, r0, n + 1)
    r0 += 1
    s = SCENARIOS[scen]
    growth_row = r0
    ws.cell(row=r0, column=1, value='Net revenue growth %').font = BLACK
    for i, v in enumerate(s['rev_growth']):
        c = ws.cell(row=r0, column=2 + i, value=v); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
    r0 += 1
    eff_row = r0
    ws.cell(row=r0, column=1, value='Efficiency ratio (Opex / Revenue)').font = BLACK
    for i, v in enumerate(s['efficiency_ratio']):
        c = ws.cell(row=r0, column=2 + i, value=v); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
    r0 += 1
    tax_row = r0
    ws.cell(row=r0, column=1, value='Effective tax rate').font = BLACK
    c = ws.cell(row=r0, column=2, value=s['tax_rate']); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
    r0 += 1
    payout_row = r0
    ws.cell(row=r0, column=1, value='Total payout ratio').font = BLACK
    c = ws.cell(row=r0, column=2, value=s['payout_ratio']); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
    r0 += 1
    bs_row = r0
    ws.cell(row=r0, column=1, value='Balance sheet / RWA growth (required equity retention)').font = BLACK
    c = ws.cell(row=r0, column=2, value=s['bs_growth']); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
    r0 += 1
    tg_row = r0
    ws.cell(row=r0, column=1, value='Terminal growth rate').font = BLACK
    c = ws.cell(row=r0, column=2, value=s['terminal_growth']); c.font = BLUE; c.number_format = PCT2; c.border = BORDER
    r0 += 1
    ke_row = r0
    ws.cell(row=r0, column=1, value='Cost of equity (CAPM)').font = BLACK
    c = ws.cell(row=r0, column=2, value=s['cost_of_equity']); c.font = BLUE; c.number_format = PCT2; c.border = BORDER
    r0 += 2
    BLOCK_ROWS[scen] = dict(growth=growth_row, eff=eff_row, tax=tax_row, payout=payout_row,
                             bs=bs_row, tg=tg_row, ke=ke_row)

# ---- CAPM build ----
ws.cell(row=r0, column=1, value='CAPM Cost of Equity Build (reference)').font = SUBHDR
ws.cell(row=r0, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=2)
r0 += 1
CAPM_RF = r0
ws.cell(row=r0, column=1, value='Risk-free rate (10Y UST, Jul 2 2026)').font = BLACK
c = ws.cell(row=r0, column=2, value=RISK_FREE_RATE); c.font = BLUE; c.number_format = PCT2; c.border = BORDER
r0 += 1
CAPM_BETA = r0
ws.cell(row=r0, column=1, value='Beta (5Y monthly, approx.)').font = BLACK
c = ws.cell(row=r0, column=2, value=GS_BETA); c.font = BLUE; c.number_format = '0.00'; c.border = BORDER
r0 += 1
CAPM_ERP = r0
ws.cell(row=r0, column=1, value='Equity risk premium (long-run US)').font = BLACK
c = ws.cell(row=r0, column=2, value=EQUITY_RISK_PREMIUM); c.font = BLUE; c.number_format = PCT1; c.border = BORDER
r0 += 1
CAPM_KE = r0
ws.cell(row=r0, column=1, value='=> Implied cost of equity (Rf + Beta x ERP)').font = BOLD
c = ws.cell(row=r0, column=2, value=f'=B{CAPM_RF}+B{CAPM_BETA}*B{CAPM_ERP}')
c.font = BLACK; c.number_format = PCT2; c.border = BORDER
r0 += 1
ws.cell(row=r0, column=1, value='(Base Case Ke input above set close to this CAPM output; Bear/Bull widen').font = NOTE
r0 += 1
ws.cell(row=r0, column=1, value='the discount rate to reflect higher/lower perceived risk in each state.)').font = NOTE
r0 += 2

# ---- Active drivers block ----
ws.cell(row=r0, column=1, value='ACTIVE DRIVERS (auto-selected from B3)').font = SUBHDR
ws.cell(row=r0, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=1 + n)
r0 += 1
ws.cell(row=r0, column=1, value='Fiscal Year')
for i, y in enumerate(years):
    ws.cell(row=r0, column=2 + i, value=f'{y}E')
style_header_row(ws, r0, n + 1)
r0 += 1
ACT_GROWTH = r0
ws.cell(row=r0, column=1, value='Net revenue growth %').font = BOLD
for i in range(n):
    col = get_column_letter(2 + i)
    f = f'=IF($B$3="Bear",{col}{BLOCK_ROWS["Bear"]["growth"]},IF($B$3="Base",{col}{BLOCK_ROWS["Base"]["growth"]},{col}{BLOCK_ROWS["Bull"]["growth"]}))'
    c = ws.cell(row=r0, column=2 + i, value=f); c.number_format = PCT1; c.border = BORDER
r0 += 1
ACT_EFF = r0
ws.cell(row=r0, column=1, value='Efficiency ratio').font = BOLD
for i in range(n):
    col = get_column_letter(2 + i)
    f = f'=IF($B$3="Bear",{col}{BLOCK_ROWS["Bear"]["eff"]},IF($B$3="Base",{col}{BLOCK_ROWS["Base"]["eff"]},{col}{BLOCK_ROWS["Bull"]["eff"]}))'
    c = ws.cell(row=r0, column=2 + i, value=f); c.number_format = PCT1; c.border = BORDER
r0 += 1
ACT_TAX = r0
ws.cell(row=r0, column=1, value='Effective tax rate').font = BOLD
f = f'=IF($B$3="Bear",B{BLOCK_ROWS["Bear"]["tax"]},IF($B$3="Base",B{BLOCK_ROWS["Base"]["tax"]},B{BLOCK_ROWS["Bull"]["tax"]}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.border = BORDER
r0 += 1
ACT_PAYOUT = r0
ws.cell(row=r0, column=1, value='Total payout ratio').font = BOLD
f = f'=IF($B$3="Bear",B{BLOCK_ROWS["Bear"]["payout"]},IF($B$3="Base",B{BLOCK_ROWS["Base"]["payout"]},B{BLOCK_ROWS["Bull"]["payout"]}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.border = BORDER
r0 += 1
ACT_BS = r0
ws.cell(row=r0, column=1, value='Balance sheet / RWA growth').font = BOLD
f = f'=IF($B$3="Bear",B{BLOCK_ROWS["Bear"]["bs"]},IF($B$3="Base",B{BLOCK_ROWS["Base"]["bs"]},B{BLOCK_ROWS["Bull"]["bs"]}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT1; c.border = BORDER
r0 += 1
ACT_TG = r0
ws.cell(row=r0, column=1, value='Terminal growth rate').font = BOLD
f = f'=IF($B$3="Bear",B{BLOCK_ROWS["Bear"]["tg"]},IF($B$3="Base",B{BLOCK_ROWS["Base"]["tg"]},B{BLOCK_ROWS["Bull"]["tg"]}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT2; c.border = BORDER
r0 += 1
ACT_KE = r0
ws.cell(row=r0, column=1, value='Cost of equity (CAPM)').font = BOLD
f = f'=IF($B$3="Bear",B{BLOCK_ROWS["Bear"]["ke"]},IF($B$3="Base",B{BLOCK_ROWS["Base"]["ke"]},B{BLOCK_ROWS["Bull"]["ke"]}))'
c = ws.cell(row=r0, column=2, value=f); c.number_format = PCT2; c.border = BORDER
r0 += 2

ws.cell(row=r0, column=1, value='Preferred dividends, held flat ($mm/yr)').font = BOLD
c = ws.cell(row=r0, column=2, value=PREFERRED_DIV_PROJECTED); c.font = BLUE; c.number_format = USD0; c.border = BORDER
ACT_PREF = r0
r0 += 1

r0 += 1
ws.cell(row=r0, column=1, value='Market data (for implied price comparison)').font = SUBHDR
ws.cell(row=r0, column=1).fill = SUBHDR_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=2)
r0 += 1
MKT_PRICE = r0
ws.cell(row=r0, column=1, value='Current share price').font = BLACK
c = ws.cell(row=r0, column=2, value=GS_PRICE); c.font = BLUE; c.number_format = USD2; c.border = BORDER
r0 += 1
MKT_SHARES = r0
ws.cell(row=r0, column=1, value='Current diluted shares outstanding (mm)').font = BLACK
c = ws.cell(row=r0, column=2, value=GS_SHARES_CURRENT_MM); c.font = BLUE; c.number_format = NUM1; c.border = BORDER
r0 += 1
MKT_CAP = r0
ws.cell(row=r0, column=1, value='=> Implied current market equity value ($mm)').font = BOLD
c = ws.cell(row=r0, column=2, value=f'=B{MKT_PRICE}*B{MKT_SHARES}'); c.number_format = USD0; c.border = BORDER
r0 += 1
BASE_EQUITY_2025 = r0
ws.cell(row=r0, column=1, value='2025A ending common equity ($mm) [BVPS x shares]').font = BLACK
c = ws.cell(row=r0, column=2, value=round(GS_HISTORY[2025]["bvps"] * GS_SHARES_HISTORY[2025], 1))
c.font = BLUE; c.number_format = USD0; c.border = BORDER
r0 += 1
BASE_REV_2025 = r0
ws.cell(row=r0, column=1, value='2025A net revenues ($mm)').font = BLACK
c = ws.cell(row=r0, column=2, value=GS_HISTORY[2025]['net_revenues']); c.font = BLUE; c.number_format = USD0; c.border = BORDER

wb.save('/home/claude/gs-valuation/models/03_DCF_and_Reverse_DCF.xlsx')
print("Assumptions sheet complete for DCF workbook.")
print(dict(ACT_GROWTH=ACT_GROWTH, ACT_EFF=ACT_EFF, ACT_TAX=ACT_TAX, ACT_PAYOUT=ACT_PAYOUT, ACT_BS=ACT_BS,
           ACT_TG=ACT_TG, ACT_KE=ACT_KE, ACT_PREF=ACT_PREF, MKT_PRICE=MKT_PRICE, MKT_SHARES=MKT_SHARES,
           MKT_CAP=MKT_CAP, BASE_EQUITY_2025=BASE_EQUITY_2025, BASE_REV_2025=BASE_REV_2025,
           CAPM_KE=CAPM_KE))
print("BLOCK_ROWS:", BLOCK_ROWS)
