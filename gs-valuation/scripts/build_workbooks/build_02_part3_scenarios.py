import sys
sys.path.insert(0, '/home/claude/gs-valuation/data')
from assumptions import *
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name='Arial', size=10, color='0000FF')
BLACK = Font(name='Arial', size=9, color='000000')
GREEN = Font(name='Arial', size=9, color='008000')
BOLD = Font(name='Arial', size=10, bold=True)
HDR = Font(name='Arial', size=9, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE = Font(name='Arial', size=14, bold=True, color='FFFFFF')
SCEN_FILL = {'Bear': PatternFill('solid', fgColor='F8CBAD'),
             'Base': PatternFill('solid', fgColor='D9E1F2'),
             'Bull': PatternFill('solid', fgColor='C6E0B4')}
NOTE = Font(name='Arial', size=8, italic=True, color='808080')
thin = Side(style='thin', color='B7B7B7')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
USD0 = '$#,##0;($#,##0);"-"'
USD2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%;(0.0%);"-"'
NUM1 = '#,##0.0'

def title_block(ws, text, span=17):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws['A1']; c.value = text; c.font = TITLE; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

wb = load_workbook('/home/claude/gs-valuation/models/02_Three_Statement_Model.xlsx')
years = PROJECTION_YEARS
n = len(years)

BEAR_GROWTH, BEAR_EFF, BEAR_TAX, BEAR_PAYOUT, BEAR_BS, BEAR_TG, BEAR_KE = 7, 8, 9, 10, 11, 12, 13
BASE_GROWTH, BASE_EFF, BASE_TAX, BASE_PAYOUT, BASE_BS, BASE_TG, BASE_KE = 17, 18, 19, 20, 21, 22, 23
BULL_GROWTH, BULL_EFF, BULL_TAX, BULL_PAYOUT, BULL_BS, BULL_TG, BULL_KE = 27, 28, 29, 30, 31, 32, 33
ACT_PREF, ACT_PB, ACT_DIVSHARE = 46, 47, 48
GROWTH_ROW = {'Bear': BEAR_GROWTH, 'Base': BASE_GROWTH, 'Bull': BULL_GROWTH}
EFF_ROW = {'Bear': BEAR_EFF, 'Base': BASE_EFF, 'Bull': BULL_EFF}
TAX_ROW = {'Bear': BEAR_TAX, 'Base': BASE_TAX, 'Bull': BULL_TAX}
PAYOUT_ROW = {'Bear': BEAR_PAYOUT, 'Base': BASE_PAYOUT, 'Bull': BULL_PAYOUT}

ws = wb.create_sheet('Scenario Summary')
title_block(ws, 'Bull / Base / Bear — Full 5-Year Projection Comparison, FY2026E-FY2030E', span=17)
ws.column_dimensions['A'].width = 30
for col_i in range(2, 18):
    ws.column_dimensions[get_column_letter(col_i)].width = 11

# Header rows: scenario band + year band
row_scen = 3
row_year = 4
col = 2
scen_start_col = {}
for scen in ['Bear', 'Base', 'Bull']:
    scen_start_col[scen] = col
    ws.merge_cells(start_row=row_scen, start_column=col, end_row=row_scen, end_column=col + n - 1)
    c = ws.cell(row=row_scen, column=col, value=f'{scen} Case')
    c.font = HDR; c.fill = SCEN_FILL[scen]; c.alignment = Alignment(horizontal='center')
    for i, y in enumerate(years):
        yc = ws.cell(row=row_year, column=col + i, value=f'{y}E')
        yc.font = HDR; yc.fill = HDR_FILL; yc.alignment = Alignment(horizontal='center')
        yc.border = BORDER
    col += n + 1  # 1 blank spacer column between scenarios

metrics = ['Net revenues ($mm)', 'Net earnings to common ($mm)', 'Ending common equity ($mm)',
           'Ending diluted shares (mm)', 'Diluted EPS ($)', 'Book value / share ($)',
           'Return on average common equity (%)']
r0 = 5
metric_row = {}
for m in metrics:
    ws.cell(row=r0, column=1, value=m).font = BOLD
    metric_row[m] = r0
    r0 += 1

base_equity_2025 = GS_HISTORY[2025]['bvps'] * GS_SHARES_HISTORY[2025]

for scen in ['Bear', 'Base', 'Bull']:
    sc = scen_start_col[scen]
    growth_r, eff_r, tax_r, payout_r = GROWTH_ROW[scen], EFF_ROW[scen], TAX_ROW[scen], PAYOUT_ROW[scen]
    for i, y in enumerate(years):
        col_letter = get_column_letter(sc + i)
        acol = get_column_letter(2 + i)  # Assumptions Bear/Base/Bull rows use B..F for 2026..2030
        prev_letter = get_column_letter(sc + i - 1) if i > 0 else None

        # Net revenues
        if i == 0:
            rev_formula = f'={base_equity_2025 / GS_HISTORY[2025]["bvps"]:.6f}'  # placeholder unused
            rev_formula = f'={GS_HISTORY[2025]["net_revenues"]}*(1+Assumptions!{acol}{growth_r})'
        else:
            rev_formula = f'={prev_letter}{metric_row["Net revenues ($mm)"]}*(1+Assumptions!{acol}{growth_r})'
        rc = ws.cell(row=metric_row['Net revenues ($mm)'], column=sc + i, value=rev_formula)
        rc.number_format = USD0; rc.border = BORDER; rc.font = BLACK

        # Net earnings to common = Rev*(1-eff)*(1-tax) - preferred
        rev_ref = f'{col_letter}{metric_row["Net revenues ($mm)"]}'
        nic_formula = (f'=({rev_ref}*(1-Assumptions!{acol}{eff_r}))*(1-Assumptions!$B${tax_r})'
                       f'-Assumptions!$B${ACT_PREF}')
        nc = ws.cell(row=metric_row['Net earnings to common ($mm)'], column=sc + i, value=nic_formula)
        nc.number_format = USD0; nc.border = BORDER; nc.font = BLACK

        # Ending equity
        nic_ref = f'{col_letter}{metric_row["Net earnings to common ($mm)"]}'
        if i == 0:
            begeq = f'{base_equity_2025:.1f}'
        else:
            begeq = f'{prev_letter}{metric_row["Ending common equity ($mm)"]}'
        endeq_formula = f'={begeq}+{nic_ref}-({nic_ref}*Assumptions!$B${payout_r})'
        ec = ws.cell(row=metric_row['Ending common equity ($mm)'], column=sc + i, value=endeq_formula)
        ec.number_format = USD0; ec.border = BORDER; ec.font = BLACK

        # Ending shares: beg shares - (buyback$ / repurchase price)
        if i == 0:
            begsh = f'{GS_SHARES_HISTORY[2025]}'
        else:
            begsh = f'{prev_letter}{metric_row["Ending diluted shares (mm)"]}'
        buyback_dollar = f'(({nic_ref}*Assumptions!$B${payout_r})*(1-Assumptions!$B${ACT_DIVSHARE}))'
        rep_price = f'(({begeq}/{begsh})*Assumptions!$B${ACT_PB})'
        endsh_formula = f'={begsh}-({buyback_dollar}/{rep_price})'
        sc_cell = ws.cell(row=metric_row['Ending diluted shares (mm)'], column=sc + i, value=endsh_formula)
        sc_cell.number_format = NUM1; sc_cell.border = BORDER; sc_cell.font = BLACK

        # EPS = NIC / average(beg,end shares)
        endsh_ref = f'{col_letter}{metric_row["Ending diluted shares (mm)"]}'
        eps_formula = f'={nic_ref}/AVERAGE({begsh},{endsh_ref})'
        epsc = ws.cell(row=metric_row['Diluted EPS ($)'], column=sc + i, value=eps_formula)
        epsc.number_format = USD2; epsc.border = BORDER; epsc.font = BLACK

        # BVPS
        endeq_ref = f'{col_letter}{metric_row["Ending common equity ($mm)"]}'
        bvps_formula = f'={endeq_ref}/{endsh_ref}'
        bvc = ws.cell(row=metric_row['Book value / share ($)'], column=sc + i, value=bvps_formula)
        bvc.number_format = USD2; bvc.border = BORDER; bvc.font = BLACK

        # ROE
        roe_formula = f'={nic_ref}/AVERAGE({begeq},{endeq_ref})'
        roec = ws.cell(row=metric_row['Return on average common equity (%)'], column=sc + i, value=roe_formula)
        roec.number_format = PCT1; roec.border = BORDER; roec.font = BLACK

r0 += 2
ws.cell(row=r0, column=1, value='Terminal-Year (2030E) Snapshot').font = BOLD
r0 += 1
ws.cell(row=r0, column=1, value='Scenario').font = HDR
ws.cell(row=r0, column=1).fill = HDR_FILL
for j, m in enumerate(['2030E EPS ($)', '2030E BVPS ($)', '2030E ROE (%)', '2030E Net Revenues ($mm)']):
    c = ws.cell(row=r0, column=2 + j, value=m); c.font = HDR; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)
r0 += 1
for scen in ['Bear', 'Base', 'Bull']:
    sc = scen_start_col[scen]
    last_col = get_column_letter(sc + n - 1)
    ws.cell(row=r0, column=1, value=scen).font = BOLD
    ws.cell(row=r0, column=1).fill = SCEN_FILL[scen]
    vals = [
        (f'={last_col}{metric_row["Diluted EPS ($)"]}', USD2),
        (f'={last_col}{metric_row["Book value / share ($)"]}', USD2),
        (f'={last_col}{metric_row["Return on average common equity (%)"]}', PCT1),
        (f'={last_col}{metric_row["Net revenues ($mm)"]}', USD0),
    ]
    for j, (f, fmt) in enumerate(vals):
        c = ws.cell(row=r0, column=2 + j, value=f); c.number_format = fmt; c.border = BORDER
    r0 += 1

r0 += 2
for line in [
    "This tab independently replicates the Bear/Base/Bull assumption chains from the Assumptions tab",
    "(rather than reading the single 'active' scenario used elsewhere), so all three cases can be",
    "compared side by side without toggling cell B3 on the Assumptions tab. Toggling B3 changes the",
    "single live scenario flowing through the Income Statement Proj. and Capital & Shares tabs, and the",
    "DCF / Reverse DCF workbook keys off that same selector for its base-case path.",
]:
    ws.cell(row=r0, column=1, value=line).font = NOTE
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=10)
    r0 += 1

order = ['Assumptions', 'Income Statement Proj.', 'Capital & Shares', 'Scenario Summary']
wb._sheets = [wb[s] for s in order]
wb.save('/home/claude/gs-valuation/models/02_Three_Statement_Model.xlsx')
print("Scenario Summary sheet complete.")
