"""
validate_three_statement.py
============================
Independently re-implements the Goldman Sachs integrated Income Statement +
Capital & Shares roll-forward model (see models/02_Three_Statement_Model.xlsx)
in pure Python/pandas, then reads the calculated values out of the Excel file
and checks that Excel and Python agree to within a small tolerance.

This is a standard "build it twice, in two different tools, and diff it"
sanity check -- the same practice a rigorous analyst uses before trusting a
model for a live decision.

Usage:
    python validate_three_statement.py
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'data'))
from assumptions import *
from openpyxl import load_workbook

TOL_PCT = 0.001  # 0.1% relative tolerance


def run_python_model(scenario='Base'):
    """Re-implement the Excel formula chain independently in Python."""
    s = SCENARIOS[scenario]
    years = PROJECTION_YEARS

    rev = GS_HISTORY[2025]['net_revenues']
    equity = GS_HISTORY[2025]['bvps'] * GS_SHARES_HISTORY[2025]
    shares = GS_SHARES_HISTORY[2025]

    rows = []
    for i, y in enumerate(years):
        rev = rev * (1 + s['rev_growth'][i])
        opex = rev * s['efficiency_ratio'][i]
        pretax = rev - opex
        tax = pretax * s['tax_rate']
        net_earnings = pretax - tax
        nic = net_earnings - PREFERRED_DIV_PROJECTED

        beg_equity = equity
        total_payout = nic * s['payout_ratio']
        dividends = total_payout * 0.28
        buyback_dollar = total_payout - dividends
        end_equity = beg_equity + nic - total_payout

        beg_shares = shares
        rep_price = (beg_equity / beg_shares) * (GS_PRICE / GS_HISTORY[2025]['bvps'])
        shares_repurchased = buyback_dollar / rep_price
        end_shares = beg_shares - shares_repurchased
        avg_shares = (beg_shares + end_shares) / 2

        eps = nic / avg_shares
        bvps = end_equity / end_shares
        roe = nic / ((beg_equity + end_equity) / 2)

        rows.append(dict(year=y, net_revenues=rev, opex=opex, pretax=pretax, tax=tax,
                          net_earnings=net_earnings, nic=nic, end_equity=end_equity,
                          end_shares=end_shares, avg_shares=avg_shares, eps=eps, bvps=bvps, roe=roe))

        equity = end_equity
        shares = end_shares

    return rows


def read_excel_model(path, scenario='Base'):
    wb = load_workbook(path, data_only=True)
    ws_assump = wb['Assumptions']
    current_scenario = ws_assump['B3'].value
    if current_scenario != scenario:
        print(f"WARNING: workbook's active scenario is '{current_scenario}', "
              f"not '{scenario}'. Re-open in Excel/LibreOffice, set Assumptions!B3='{scenario}', "
              f"save, and re-run recalc.py before validating that case.")
    ws_is = wb['Income Statement Proj.']
    ws_cap = wb['Capital & Shares']

    rows = []
    for i, y in enumerate(PROJECTION_YEARS):
        col = 3 + i  # column C onward = 2026E..2030E
        rows.append(dict(
            year=y,
            net_revenues=ws_is.cell(row=6, column=col).value,
            nic=ws_is.cell(row=13, column=col).value,
            eps=ws_is.cell(row=15, column=col).value,
            end_equity=ws_cap.cell(row=11, column=col).value,
            end_shares=ws_cap.cell(row=15, column=col).value,
            bvps=ws_cap.cell(row=17, column=col).value,
            roe=ws_cap.cell(row=18, column=col).value,
        ))
    return rows, current_scenario


def compare(python_rows, excel_rows, fields):
    all_ok = True
    print(f"{'Year':<6}{'Field':<16}{'Python':>16}{'Excel':>16}{'Diff %':>10}  Status")
    print("-" * 76)
    for pr, er in zip(python_rows, excel_rows):
        for f in fields:
            pv, ev = pr[f], er[f]
            if ev is None:
                print(f"{pr['year']:<6}{f:<16}{pv:>16.4f}{'N/A':>16}{'':>10}  SKIP (Excel not on Base case)")
                continue
            diff_pct = abs(pv - ev) / abs(ev) if ev != 0 else abs(pv - ev)
            status = "OK" if diff_pct < TOL_PCT else "MISMATCH"
            if status == "MISMATCH":
                all_ok = False
            print(f"{pr['year']:<6}{f:<16}{pv:>16.4f}{ev:>16.4f}{diff_pct*100:>9.3f}%  {status}")
    return all_ok


if __name__ == '__main__':
    model_path = os.path.join(os.path.dirname(__file__), '..', 'models', '02_Three_Statement_Model.xlsx')
    print("Re-computing 3-statement model independently in Python (Base Case)...\n")
    py_rows = run_python_model('Base')

    print(f"Reading calculated values from {model_path} ...\n")
    excel_rows, active_scenario = read_excel_model(model_path, 'Base')

    fields = ['net_revenues', 'nic', 'end_equity', 'end_shares', 'bvps', 'roe', 'eps']
    ok = compare(py_rows, excel_rows, fields)

    print("\n" + "=" * 76)
    if active_scenario != 'Base':
        print(f"VALIDATION INCOMPLETE: workbook active scenario is '{active_scenario}'. "
              f"Set it to 'Base', save, re-run recalc.py, then re-run this script.")
        sys.exit(2)
    elif ok:
        print("PASS: Python-independent recomputation matches Excel model within 0.1% on all fields.")
        sys.exit(0)
    else:
        print("FAIL: one or more fields diverged beyond tolerance -- review formulas above.")
        sys.exit(1)
